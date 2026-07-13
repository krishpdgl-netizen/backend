from fastapi import FastAPI, UploadFile, File as FastAPIFile
from sqlalchemy import create_engine
from sqlalchemy import text
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional 
from excel_helper import (
    add_projection,
    get_sales,
    update_achieved,
    get_all_weeks,
    get_sales_for_employees,
    admin_update_projection,
    raise_change_request,
    get_change_requests,
    resolve_change_request,
    current_week,
    FILE_NAME
)
from datetime import datetime
from zoneinfo import ZoneInfo
from datetime import date, timedelta
import json
import secrets
from pydantic import BaseModel
from typing import List
from sqlalchemy import text
class MeetingRequest(BaseModel):
    title: str
    description: str = ""
    meeting_date: str
    start_slot: int
    end_slot: int
    organizer_id: int
    location: str = ""
    attendees: List[int]
import requests
from pywebpush import webpush, WebPushException
import os
from pydantic import BaseModel
from typing import Optional

class LeaveRequest(BaseModel):
    user_id: int
    employee_name: str
    leave_type: str
    leave_type_id: Optional[int] = None
    start_date: str
    end_date: str
    reason: str

class TravelRequestIn(BaseModel):
    user_id: int
    employee_name: str
    origin: str
    destination: str
    travel_mode: str = "Flight"
    start_date: str
    end_date: str
    purpose: str
    estimated_cost: Optional[float] = None

class TravelMeetingIn(BaseModel):
    meeting_date: str
    start_slot: int
    end_slot: int
    location: str = ""
    
app = FastAPI()

from typing import Optional
from pydantic import BaseModel

class EmployeeVoiceRequest(BaseModel):
    user_id: int
    employee_name: str
    is_anonymous: bool
    category: str
    priority: str
    subject: str
    description: str
    attachment: Optional[str] = None

from pydantic import BaseModel
from typing import Optional
 
class VoidRequest(BaseModel):
    void_reason: str



# ── AUTO-CREATE TABLES ON STARTUP ────────────────────────
@app.on_event("startup")
def create_tables():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS calendar_tasks (
                id          SERIAL PRIMARY KEY,
                user_id     INT NOT NULL,
                title       TEXT NOT NULL,
                description TEXT DEFAULT '',
                task_date   DATE NOT NULL,
                start_slot  INT NOT NULL,
                end_slot    INT NOT NULL,
                color       TEXT DEFAULT '#7c3aed',
                status      TEXT DEFAULT 'pending',
                created_at  TIMESTAMP DEFAULT NOW()
            )
        """))
        # Patch any columns missing from an older version of the table
        for col_sql in [
            "ALTER TABLE calendar_tasks ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'pending'",
            "ALTER TABLE calendar_tasks ADD COLUMN IF NOT EXISTS description TEXT DEFAULT ''",
            "ALTER TABLE calendar_tasks ADD COLUMN IF NOT EXISTS color TEXT DEFAULT '#7c3aed'",
            "ALTER TABLE calendar_tasks ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT NOW()",
        ]:
            conn.execute(text(col_sql))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS meetings (
                id           SERIAL PRIMARY KEY,
                title        TEXT NOT NULL,
                description  TEXT DEFAULT '',
                organizer_id INT NOT NULL,
                meeting_date DATE NOT NULL,
                start_slot   INT NOT NULL,
                end_slot     INT NOT NULL,
                location     TEXT DEFAULT '',
                created_at   TIMESTAMP DEFAULT NOW()
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS meeting_attendees (
                id              SERIAL PRIMARY KEY,
                meeting_id      INT NOT NULL,
                user_id         INT NOT NULL,
                response_status TEXT DEFAULT 'pending'
            )
        """))
        # ── team_requests: pending/approved/rejected join requests ──
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS team_requests (
                id          SERIAL PRIMARY KEY,
                manager_id  INT NOT NULL,
                employee_id INT NOT NULL,
                status      TEXT DEFAULT 'pending',
                created_at  TIMESTAMP DEFAULT NOW()
            )
        """))
        # ── tasks: file attachment columns (idempotent) ─────────────
        conn.execute(text(
            "ALTER TABLE tasks ADD COLUMN IF NOT EXISTS file_url  TEXT"
        ))
        conn.execute(text(
            "ALTER TABLE tasks ADD COLUMN IF NOT EXISTS file_name TEXT"
        ))

    # ── leave_requests / team_members migrations ─────────────────────
    # Run in their OWN transaction, isolated from the block above, and
    # wrapped in try/except so that if anything here doesn't match an
    # existing schema, it can never prevent the app from starting up
    # (which is what previously broke EVERY endpoint, surfacing as a
    # misleading CORS error in the browser).
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS leave_requests (
                    id            SERIAL PRIMARY KEY,
                    user_id       INT NOT NULL,
                    employee_name TEXT NOT NULL,
                    leave_type    TEXT NOT NULL,
                    start_date    DATE NOT NULL,
                    end_date      DATE NOT NULL,
                    reason        TEXT DEFAULT '',
                    status        TEXT DEFAULT 'Pending',
                    created_at    TIMESTAMP DEFAULT NOW(),
                    approved_at   TIMESTAMP
                )
            """))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS team_members (
                    id          SERIAL PRIMARY KEY,
                    manager_id  INT NOT NULL,
                    employee_id INT NOT NULL
                )
            """))
            # approver_id -> who currently needs to act on this request
            #                (the requester's manager, or an admin if the
            #                requester IS a manager / has no manager / was
            #                escalated)
            # escalated   -> true once a manager has forwarded an
            #                employee's request up to admin
            conn.execute(text(
                "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS approver_id INT"
            ))
            conn.execute(text(
                "ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS escalated BOOLEAN DEFAULT FALSE"
            ))
            # Backfill: some approved team_requests rows may never have been
            # mirrored into team_members (e.g. from before this safeguard
            # existed). This keeps manager<->employee routing correct
            # everywhere that reads team_members, not just leave requests.
            conn.execute(text("""
                INSERT INTO team_members (manager_id, employee_id)
                SELECT tr.manager_id, tr.employee_id
                FROM team_requests tr
                WHERE tr.status = 'approved'
                AND NOT EXISTS (
                    SELECT 1 FROM team_members tm
                    WHERE tm.manager_id = tr.manager_id
                    AND   tm.employee_id = tr.employee_id
                )
            """))
            # Re-route any still-pending leave requests whose approver_id
            # currently points at the requester's own manager mismatch
            # (e.g. were previously mis-routed straight to admin before the
            # team_members backfill above ran). Only touches requests from
            # employees (not managers/admins) so manager->admin escalation
            # behavior is left untouched.
            pending_employee_requests = conn.execute(text("""
                SELECT lr.id, lr.user_id
                FROM leave_requests lr
                JOIN users u ON u.id = lr.user_id
                WHERE lr.status = 'Pending'
                AND   lr.escalated IS NOT TRUE
                AND   LOWER(u.role) = 'employee'
            """)).fetchall()
            for lr_id, lr_user_id in pending_employee_requests:
                mgr_row = conn.execute(text("""
                    SELECT manager_id FROM team_members
                    WHERE employee_id = :uid
                    ORDER BY id DESC LIMIT 1
                """), {"uid": lr_user_id}).first()
                if mgr_row and mgr_row[0]:
                    conn.execute(text("""
                        UPDATE leave_requests
                        SET approver_id = :mgr_id
                        WHERE id = :lr_id
                    """), {"mgr_id": mgr_row[0], "lr_id": lr_id})
    except Exception as e:
        print(f"[startup migration warning] leave_requests/team_members migration failed: {e}")

    # ── password reset / push notifications / leave balances ─────────
    # Own isolated try/except, same reasoning as above: a problem here
    # must never prevent the whole app from starting.
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS password_reset_tokens (
                    id          SERIAL PRIMARY KEY,
                    user_id     INT NOT NULL,
                    token       TEXT NOT NULL UNIQUE,
                    expires_at  TIMESTAMP NOT NULL,
                    used        BOOLEAN DEFAULT FALSE,
                    created_at  TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS push_subscriptions (
                    id          SERIAL PRIMARY KEY,
                    user_id     INT NOT NULL,
                    endpoint    TEXT NOT NULL UNIQUE,
                    p256dh      TEXT NOT NULL,
                    auth        TEXT NOT NULL,
                    created_at  TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS leave_types (
                    id                 SERIAL PRIMARY KEY,
                    name               TEXT NOT NULL UNIQUE,
                    annual_quota       NUMERIC,             -- NULL = not balance-tracked (e.g. WFH)
                    carry_forward      BOOLEAN DEFAULT FALSE,
                    is_balance_tracked BOOLEAN DEFAULT TRUE,
                    created_at         TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS leave_balances (
                    id            SERIAL PRIMARY KEY,
                    user_id       INT NOT NULL,
                    leave_type_id INT NOT NULL REFERENCES leave_types(id),
                    year          INT NOT NULL,
                    allocated     NUMERIC NOT NULL DEFAULT 0,
                    used          NUMERIC NOT NULL DEFAULT 0,
                    created_at    TIMESTAMP DEFAULT NOW(),
                    UNIQUE (user_id, leave_type_id, year)
                )
            """))
            # Link leave_requests to a real leave type + day count, without
            # touching the existing free-text leave_type column (old rows
            # keep working, nothing already stored gets orphaned).
            conn.execute(text("ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS leave_type_id INT"))
            conn.execute(text("ALTER TABLE leave_requests ADD COLUMN IF NOT EXISTS days NUMERIC"))

            # Seed sensible default leave types, once.
            existing_types = conn.execute(text("SELECT COUNT(*) FROM leave_types")).scalar()
            if existing_types == 0:
                conn.execute(text("""
                    INSERT INTO leave_types (name, annual_quota, carry_forward, is_balance_tracked) VALUES
                    ('Casual Leave', 12, FALSE, TRUE),
                    ('Sick Leave', 8, FALSE, TRUE),
                    ('Earned Leave', 15, TRUE, TRUE),
                    ('Maternity / Paternity Leave', 90, FALSE, TRUE),
                    ('Work From Home', NULL, FALSE, FALSE),
                    ('Half Day', NULL, FALSE, FALSE)
                """))
    except Exception as e:
        print(f"[startup migration warning] password reset / push / leave balance migration failed: {e}")

    # ── one-time: hash any plaintext passwords still in the DB ────────
    # bcrypt hashes always start with $2b$/$2a$/$2y$, so any row whose
    # password doesn't start with that is still plaintext from before
    # this migration existed. We hash it in place, once, automatically.
    try:
        with engine.begin() as conn:
            rows = conn.execute(text("SELECT id, password FROM users")).mappings().all()
            for row in rows:
                pw = row["password"]
                if pw and not pw.startswith(("$2a$", "$2b$", "$2y$")):
                    conn.execute(
                        text("UPDATE users SET password=:pw WHERE id=:id"),
                        {"pw": hash_password(pw), "id": row["id"]}
                    )
    except Exception as e:
        print(f"[startup migration warning] password hashing migration failed: {e}")

    # ── print tracking (Print Center) ─────────────────────────────
    # Own isolated try/except, same reasoning as above.
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS print_logs (
                    id              SERIAL PRIMARY KEY,
                    user_id         INT NOT NULL,
                    user_name       TEXT NOT NULL,
                    document_title  TEXT DEFAULT '',
                    printed_for     TEXT NOT NULL,
                    given_to        TEXT NOT NULL,
                    printed_at      TIMESTAMP DEFAULT NOW()
                )
            """))
    except Exception as e:
        print(f"[startup migration warning] print_logs migration failed: {e}")

    # ── travel_requests (Business Travel Requests) ────────────────
    # Own isolated try/except, same reasoning as above.
    try:
        with engine.begin() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS travel_requests (
                    id                SERIAL PRIMARY KEY,
                    user_id           INT NOT NULL,
                    employee_name     TEXT NOT NULL,
                    origin            TEXT NOT NULL,
                    destination       TEXT NOT NULL,
                    travel_mode       TEXT DEFAULT 'Flight',
                    start_date        DATE NOT NULL,
                    end_date          DATE NOT NULL,
                    purpose           TEXT DEFAULT '',
                    estimated_cost    NUMERIC,
                    status            TEXT DEFAULT 'Pending',
                    meeting_id        INT,
                    created_at        TIMESTAMP DEFAULT NOW(),
                    approved_at       TIMESTAMP,
                    last_reminder_at  TIMESTAMP DEFAULT NOW()
                )
            """))
            for col_sql in [
                "ALTER TABLE travel_requests ADD COLUMN IF NOT EXISTS meeting_id INT",
                "ALTER TABLE travel_requests ADD COLUMN IF NOT EXISTS last_reminder_at TIMESTAMP DEFAULT NOW()",
            ]:
                conn.execute(text(col_sql))
    except Exception as e:
        print(f"[startup migration warning] travel_requests migration failed: {e}")


# ── CORS ─────────────────────────────────────────────────
# allow_origins=["*"] with allow_credentials=False is correct.
# We also add an explicit OPTIONS catch-all so Railway's proxy
# never swallows the preflight before FastAPI can respond to it.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
    expose_headers=["*"],
    max_age=600,
)

from fastapi import Request
from fastapi.responses import Response

@app.options("/{rest_of_path:path}")
async def preflight_handler(rest_of_path: str, request: Request):
    return Response(
        status_code=200,
        headers={
            "Access-Control-Allow-Origin":  "*",
            "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS, PATCH",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Max-Age":       "600",
        },
    )
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://neondb_owner:npg_ceN7gVxPlpA5@ep-rough-wind-atil7hvd-pooler.c-9.us-east-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require"
)

engine = create_engine(
    DATABASE_URL,
    connect_args={"sslmode":"require"}
)

# ================================================================
# AUTH -- signed session tokens + role guards
# ================================================================
import hmac, hashlib, base64
import json as _auth_json
import time as _auth_time
import bcrypt
from fastapi import Header, Depends, HTTPException

AUTH_SECRET_KEY = os.getenv("AUTH_SECRET_KEY", "panache-dev-secret-CHANGE-ME-IN-PRODUCTION")
TOKEN_VALID_HOURS = 12


def hash_password(plain: str) -> str:
    """Hash a plaintext password for storage. Never store plain text."""
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, stored: str) -> bool:
    """Check a plaintext password against a stored bcrypt hash."""
    if not stored:
        return False
    try:
        return bcrypt.checkpw(plain.encode("utf-8"), stored.encode("utf-8"))
    except (ValueError, TypeError):
        # stored value isn't a valid bcrypt hash (shouldn't happen after
        # migration, but fail closed rather than crash).
        return False


def _create_token(user_id: int, role: str) -> str:
    payload = {
        "uid": user_id,
        "role": role,
        "exp": int(_auth_time.time()) + TOKEN_VALID_HOURS * 3600,
    }
    raw = _auth_json.dumps(payload, separators=(",", ":")).encode()
    payload_b64 = base64.urlsafe_b64encode(raw).decode().rstrip("=")
    sig = hmac.new(AUTH_SECRET_KEY.encode(), payload_b64.encode(), hashlib.sha256).hexdigest()
    return f"{payload_b64}.{sig}"


def _verify_token(token: str):
    try:
        payload_b64, sig = token.split(".", 1)
        expected_sig = hmac.new(AUTH_SECRET_KEY.encode(), payload_b64.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, expected_sig):
            return None
        padded = payload_b64 + "=" * (-len(payload_b64) % 4)
        payload = _auth_json.loads(base64.urlsafe_b64decode(padded))
        if payload.get("exp", 0) < _auth_time.time():
            return None
        return payload
    except Exception:
        return None


def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Not logged in. Please log in again.")
    token = authorization.split(" ", 1)[1].strip()
    payload = _verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Session expired or invalid. Please log in again.")
    return payload


def require_roles(*roles):
    def _checker(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="You don't have permission to do that.")
        return user
    return _checker


def _ist_today():
    return datetime.now(ZoneInfo("Asia/Kolkata")).date()


def slot_to_time(slot):

    hour = slot // 2
    minute = (slot % 2) * 30

    return f"{hour:02}:{minute:02}"

@app.get("/")
def home():
    return {"message":"Panache API Running"}

@app.post("/create-user")
def create_user(_admin: dict = Depends(require_roles("admin"))):

    with engine.connect() as conn:

        conn.execute(
            text("""
            INSERT INTO users
            (full_name,email,password,role)
            VALUES
            (
                'Admin User',
                'admin@panache.com',
                :password,
                'admin'
            )
            """),
            {"password": hash_password("admin123")}
        )

        conn.commit()

    return {"message":"User Created"}


class LoginRequest(BaseModel):
    email: str
    password: str


@app.post("/login")
def login(data: LoginRequest):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT *
                FROM users
                WHERE email=:email
            """),
            {
                "email": data.email
            }
        )

        user = result.fetchone()

    if user and verify_password(data.password, user.password):
        token = _create_token(user.id, user.role)
        return {
    "success": True,
    "id": user.id,
    "name": user.full_name,
    "email": user.email,
    "role": user.role,
    "token": token
     
    }

    return {
        "success": False
    }


# ── PASSWORD RESET ─────────────────────────────────────────
# Frontend origin the reset link points back to.
FRONTEND_URL = "https://panache-workforce-management.vercel.app"


@app.post("/forgot-password")
def forgot_password(email: str):
    user = None
    token = None

    with engine.begin() as conn:
        user = conn.execute(
            text("SELECT id, full_name FROM users WHERE email=:email"), {"email": email}
        ).mappings().first()

        if user:
            token = secrets.token_urlsafe(32)
            expires = datetime.utcnow() + timedelta(minutes=30)
            conn.execute(
                text("""
                    INSERT INTO password_reset_tokens (user_id, token, expires_at)
                    VALUES (:uid, :token, :exp)
                """),
                {"uid": user["id"], "token": token, "exp": expires}
            )

    # Send the email outside the transaction. Always return the SAME
    # generic message regardless of whether the account existed, so this
    # endpoint can't be used to check which emails are registered.
    if user:
        try:
            reset_link = f"{FRONTEND_URL}/reset-password.html?token={token}"
            send_email(
                email,
                "Reset your Panache WMS password",
                f"""
                <h2>Password Reset Requested</h2>
                <p>Hi {user['full_name']},</p>
                <p>Click the link below to set a new password. This link expires in 30 minutes.</p>
                <p><a href="{reset_link}">{reset_link}</a></p>
                <p>If you didn't request this, you can safely ignore this email — your password won't change.</p>
                <p>Regards,<br>Panache WMS</p>
                """
            )
        except Exception as e:
            print(f"[forgot-password email warning] {e}")

    return {
        "success": True,
        "message": "If that email is registered, a reset link has been sent."
    }


@app.get("/verify-reset-token")
def verify_reset_token(token: str):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT expires_at, used FROM password_reset_tokens WHERE token=:token"),
            {"token": token}
        ).mappings().first()

    if not row:
        return {"valid": False, "reason": "not_found"}
    if row["used"]:
        return {"valid": False, "reason": "already_used"}
    if row["expires_at"] < datetime.utcnow():
        return {"valid": False, "reason": "expired"}
    return {"valid": True}


class ResetPasswordIn(BaseModel):
    token: str
    new_password: str


@app.post("/reset-password")
def reset_password(data: ResetPasswordIn):
    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT id, user_id, expires_at, used FROM password_reset_tokens WHERE token=:token"),
            {"token": data.token}
        ).mappings().first()

        if not row:
            return {"success": False, "message": "Invalid or expired reset link."}
        if row["used"]:
            return {"success": False, "message": "This reset link has already been used."}
        if row["expires_at"] < datetime.utcnow():
            return {"success": False, "message": "This reset link has expired. Please request a new one."}
        if len(data.new_password) < 6:
            return {"success": False, "message": "Password must be at least 6 characters."}

        conn.execute(text("UPDATE users SET password=:pw WHERE id=:uid"), {"pw": hash_password(data.new_password), "uid": row["user_id"]})
        conn.execute(text("UPDATE password_reset_tokens SET used=TRUE WHERE id=:id"), {"id": row["id"]})

    return {"success": True, "message": "Password updated successfully. You can now log in."}


class RegisterRequest(BaseModel):
    fullname: str
    email: str
    password: str
    role: str


@app.post("/register")
def register_user(data: RegisterRequest, _admin: dict = Depends(require_roles("admin"))):

    with engine.connect() as conn:

        existing = conn.execute(
            text("""
                SELECT id
                FROM users
                WHERE full_name = :full_name
            """),
            {
                "full_name": data.fullname
            }
        ).fetchone()

        if existing:
            return {
                "success": False,
                "message": "Username already exists"
            }

        conn.execute(
            text("""
                INSERT INTO users
                (full_name,email,password,role)
                VALUES
                (:full_name,:email,:password,:role)
            """),
            {
                "full_name": data.fullname,
                "email": data.email,
                "password": hash_password(data.password),
                "role": data.role
            }
        )

        conn.commit()

    return {
        "success": True,
        "message": "User created"
    }

@app.get("/profile")
def get_profile(user_id: int):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT *
                FROM users
                WHERE id = :id
            """),
            {"id": user_id}
        )

        user = result.fetchone()

    if not user:
        return {"success": False}

    return {
        "success": True,
        "id": user.id,
        "name": user.full_name,
        "email": user.email,
        "role": user.role
    }


@app.get("/user")
def get_user(email: str):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT full_name,
                       email,
                       role
                FROM users
                WHERE email = :email
            """),
            {
                "email": email
            }
        )

        user = result.fetchone()

    if not user:
        return {
            "success": False
        }

    return {
        "success": True,
        "full_name": user.full_name,
        "email": user.email,
        "role": user.role
    }

@app.post("/create-task")
def create_task(
    title: str,
    description: str,
    assigned_to: int,
    assigned_by: int,
    priority: str,
    status: str,
    due_date: str
):

    with engine.connect() as conn:

        conn.execute(
            text("""
                INSERT INTO tasks
                (
                    title,
                    description,
                    assigned_to,
                    assigned_by,
                    priority,
                    status,
                    due_date
                )
                VALUES
                (
                    :title,
                    :description,
                    :assigned_to,
                    :assigned_by,
                    :priority,
                    :status,
                    :due_date
                )
            """),
            {
                "title": title,
                "description": description,
                "assigned_to": assigned_to,
                "assigned_by": assigned_by,
                "priority": priority,
                "status": status,
                "due_date": due_date
            }
        )

        conn.commit()

    try:
        send_push(assigned_to, "New Task Assigned 📋", f"{title} — due {due_date}")
    except Exception as e:
        print(f"[push warning] {e}")

    return {
        "success": True,
        "message": "Task created"
    }

@app.get("/my-tasks")
def my_tasks(user_id: int):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT *
                FROM tasks
                WHERE assigned_to = :user_id
                AND status != 'Completed'
                ORDER BY created_at DESC
            """),
            {
                "user_id": user_id
            }
        )

        tasks = [
            dict(row._mapping)
            for row in result
        ]

    return tasks


@app.post("/cron/task-due-reminders")
def task_due_reminders():
    """
    Pushes one reminder per employee for tasks due TODAY that aren't done
    yet. Not triggered automatically — schedule an external cron (Railway
    Cron, cron-job.org, GitHub Actions, etc) to POST here once a day.
    """
    today = date.today().isoformat()
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT assigned_to, COUNT(*) AS cnt
                FROM tasks
                WHERE due_date = :today
                AND status NOT IN ('Completed', 'Pending Review')
                GROUP BY assigned_to
            """),
            {"today": today}
        ).mappings().all()

    sent = 0
    for row in rows:
        try:
            n = row["cnt"]
            send_push(
                row["assigned_to"],
                "Task Due Today ⏰",
                f"You have {n} task{'s' if n != 1 else ''} due today."
            )
            sent += 1
        except Exception as e:
            print(f"[push warning] {e}")

    return {"success": True, "reminders_sent": sent}


@app.put("/update-task-status")
def update_task_status(
    task_id: int,
    status: str
):

    with engine.connect() as conn:

        conn.execute(
            text("""
                UPDATE tasks
                SET status = :status
                WHERE id = :task_id
            """),
            {
                "status": status,
                "task_id": task_id
            }
        )

        conn.commit()

    return {
        "success": True
    }


@app.get("/dashboard-stats")
def dashboard_stats():

    with engine.connect() as conn:

        total_users = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM users
            """)
        ).scalar()

        total_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
            """)
        ).scalar()

        pending_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE Lower(status) = 'pending'
            """)
        ).scalar()

        completed_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE LOWER(status) = 'complete'
            """)
        ).scalar()

    return {
        "total_users": total_users,
        "total_tasks": total_tasks,
        "pending_tasks": pending_tasks,
        "completed_tasks": completed_tasks
    }

@app.post("/complete-task")
def complete_task(task_id:int):

    with engine.connect() as conn:

        task = conn.execute(
            text("""
                SELECT *
                FROM tasks
                WHERE id=:id
            """),
            {"id":task_id}
        ).fetchone()

        conn.execute(
            text("""
                UPDATE tasks
                SET status='Completed'
                WHERE id=:id
            """),
            {"id":task_id}
        )

        conn.execute(
    text("""
        INSERT INTO task_history
        (
            task_id,
            employee_id,
            task_title,
            submitted_at
        )
        VALUES
        (
            :task_id,
            :employee_id,
            :task_title,
            :submitted_at
        )
    """),
    {
        "task_id": task.id,
        "employee_id": task.assigned_to,
        "task_title": task.title,
        "submitted_at": datetime.now(
            ZoneInfo("Asia/Kolkata")
        ).isoformat()
    }
)
        conn.commit()

    return {"success":True}

@app.get("/history")
def history(user_id:int):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT *
                FROM task_history
                WHERE employee_id=:id
                ORDER BY submitted_at DESC
            """),
            {"id":user_id}
        )

        history = [
            dict(row._mapping)
            for row in result
        ]

    return history


@app.get("/productivity-score")
def productivity_score(user_id:int):

    with engine.connect() as conn:

        total = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to=:id
            """),
            {"id":user_id}
        ).scalar()

        completed = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to=:id
                AND status='Completed'
            """),
            {"id":user_id}
        ).scalar()

    score = 0

    if total > 0:
        score = round((completed/total)*100)

    return {
        "total_tasks": total,
        "completed_tasks": completed,
        "score": score
    }

@app.get("/employees")
def get_employees():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT
                id,
                full_name,
                email,
                role
                FROM users
                WHERE role IN ('employee', 'intern')
            """)
        )

        employees = [
            dict(row._mapping)
            for row in result
        ]

    return employees

@app.get("/all-tasks")
def get_all_tasks():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT *
                FROM tasks
                ORDER BY id DESC
            """)
        )

        tasks = [
            dict(row._mapping)
            for row in result
        ]

    return tasks



@app.post("/delete-task")
def delete_task(task_id:int, _staff: dict = Depends(require_roles("admin", "manager"))):

    with engine.connect() as conn:

        conn.execute(
            text("""
                DELETE
                FROM tasks
                WHERE id=:id
            """),
            {"id":task_id}
        )

        conn.commit()

    return {
        "success":True
    }

@app.get("/employee-performance")
def employee_performance():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
            SELECT
            u.id,
            u.full_name,

            COUNT(t.id) AS assigned,

            COUNT(
                CASE
                WHEN t.status='Completed'
                THEN 1
                END
            ) AS completed

            FROM users u

            LEFT JOIN tasks t
            ON u.id=t.assigned_to

            WHERE u.role IN ('employee', 'intern')

            GROUP BY u.id,u.full_name
            """)
        )

        data=[]

        for row in result:

            assigned=row.assigned
            completed=row.completed

            score=0

            if assigned>0:
                score=round((completed/assigned)*100)

            data.append({
                "id":row.id,
                "name":row.full_name,
                "assigned":assigned,
                "completed":completed,
                "score":score
            })

    return data
@app.post("/edit-task")
def edit_task(
    task_id:int,
    title:str,
    description:str,
    assigned_to:int,
    priority:str,
    due_date:str
):

    with engine.connect() as conn:

        conn.execute(
            text("""
                UPDATE tasks
                SET
                    title=:title,
                    description=:description,
                    assigned_to=:assigned_to,
                    priority=:priority,
                    due_date=:due_date
                WHERE id=:task_id
            """),
            {
                "task_id": task_id,
                "title": title,
                "description": description,
                "assigned_to": assigned_to,
                "priority": priority,
                "due_date": due_date
            }
        )

        conn.commit()

    return {
        "success": True
    }

@app.get("/task")
def get_task(task_id:int):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT *
                FROM tasks
                WHERE id=:id
            """),
            {"id":task_id}
        )

        task = result.fetchone()

    if not task:
        return {"success":False}

    return dict(task._mapping)

@app.get("/recent-activity")
def recent_activity():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT
                task_id,
                employee_id,
                task_title,
                completed_at
                FROM task_history
                ORDER BY completed_at DESC
                LIMIT 20
            """)
        )

        activity = [
            dict(row._mapping)
            for row in result
        ]

    return activity

@app.get("/ceo/overview")
def ceo_overview(_staff: dict = Depends(require_roles("ceo", "admin"))):
    """
    Single aggregated payload for the CEO/management dashboard.
    Read-only — pulls headcount, attendance, payroll, sales, leave,
    travel and task numbers across the whole company in one call.
    """
    today = _ist_today().isoformat()
    month_start = _ist_today().replace(day=1).isoformat()
    payroll_month = _ist_today().strftime("%Y-%m")

    with engine.connect() as conn:

        # ── headcount by role ─────────────────────────────
        headcount_rows = conn.execute(
            text("SELECT role, COUNT(*) AS cnt FROM users GROUP BY role")
        ).mappings().all()
        headcount = {r["role"]: r["cnt"] for r in headcount_rows}
        total_headcount = sum(headcount.values())

        # ── attendance today ──────────────────────────────
        att_today_rows = conn.execute(
            text("SELECT status, COUNT(*) AS cnt FROM attendance WHERE att_date=:d GROUP BY status"),
            {"d": today}
        ).mappings().all()
        att_today = {r["status"]: r["cnt"] for r in att_today_rows}
        present_today = att_today.get("Present", 0) + att_today.get("Work From Home", 0)

        # ── attendance this month (company-wide %) ────────
        month_att = conn.execute(
            text("""
                SELECT
                    COUNT(*) FILTER (WHERE status IN ('Present','Work From Home')) AS present,
                    COUNT(*) AS total
                FROM attendance
                WHERE att_date >= :m
            """),
            {"m": month_start}
        ).mappings().first()
        attendance_pct = round((month_att["present"] / month_att["total"]) * 100, 1) if month_att and month_att["total"] else 0

        # ── payroll cost, latest generated month ──────────
        payroll_row = conn.execute(
            text("""
                SELECT payroll_month, SUM(net_salary) AS total_cost, COUNT(*) AS emp_count,
                       BOOL_AND(status='Locked') AS all_locked
                FROM payroll
                WHERE payroll_month = (SELECT MAX(payroll_month) FROM payroll)
                GROUP BY payroll_month
            """)
        ).mappings().first()

        # ── sales this week ────────────────────────────────
        try:
            all_emp_ids = [row["id"] for row in conn.execute(text("SELECT id FROM users")).mappings().all()]
            this_week = current_week()
            sales_rows = get_sales_for_employees(all_emp_ids, this_week)
            sales_projected = sum(float(r.get("projected") or 0) for r in sales_rows) if sales_rows else 0
            sales_achieved  = sum(float(r.get("achieved") or 0) for r in sales_rows) if sales_rows else 0
        except Exception:
            sales_projected, sales_achieved = 0, 0

        # ── leave: pending + approved this month ──────────
        leave_row = conn.execute(
            text("""
                SELECT
                    COUNT(*) FILTER (WHERE status='Pending') AS pending,
                    COUNT(*) FILTER (WHERE status='Approved' AND start_date >= :m) AS approved_this_month
                FROM leave_requests
            """),
            {"m": month_start}
        ).mappings().first()

        # ── travel: pending, flagged if estimated cost is high ─
        travel_pending = conn.execute(
            text("""
                SELECT id, employee_name, destination, estimated_cost, start_date
                FROM travel_requests
                WHERE status='Pending'
                ORDER BY estimated_cost DESC NULLS LAST
                LIMIT 10
            """)
        ).mappings().all()
        high_value_threshold = 25000  # tune to your org's approval policy
        travel_high_value = [dict(r) for r in travel_pending if (r["estimated_cost"] or 0) >= high_value_threshold]

        # ── tasks: company-wide completion rate ────────────
        task_row = conn.execute(
            text("""
                SELECT
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE status='Completed') AS completed
                FROM tasks
            """)
        ).mappings().first()
        task_completion_pct = round((task_row["completed"] / task_row["total"]) * 100, 1) if task_row and task_row["total"] else 0

    return {
        "headcount": {"total": total_headcount, "by_role": headcount},
        "attendance_today": {"present": present_today, "total": total_headcount, "breakdown": dict(att_today)},
        "attendance_month_pct": attendance_pct,
        "payroll": {
            "month": payroll_row["payroll_month"] if payroll_row else payroll_month,
            "total_cost": float(payroll_row["total_cost"]) if payroll_row and payroll_row["total_cost"] else 0,
            "employee_count": payroll_row["emp_count"] if payroll_row else 0,
            "locked": bool(payroll_row["all_locked"]) if payroll_row else False,
        },
        "sales": {"projected_this_week": sales_projected, "achieved_this_week": sales_achieved},
        "leave": {"pending": leave_row["pending"] if leave_row else 0, "approved_this_month": leave_row["approved_this_month"] if leave_row else 0},
        "travel": {"pending_count": len(travel_pending), "high_value_pending": travel_high_value, "threshold": high_value_threshold},
        "tasks": {"total": task_row["total"] if task_row else 0, "completed": task_row["completed"] if task_row else 0, "completion_pct": task_completion_pct},
    }


@app.get("/dashboard-summary")
def dashboard_summary():

    with engine.connect() as conn:

        employees = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM users
                WHERE role='employee'
            """)
        ).scalar()

        tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
            """)
        ).scalar()

        completed = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE status='Completed'
            """)
        ).scalar()

    productivity = 0

    if tasks > 0:
        productivity = round(completed/tasks*100)

    return {
        "employees": employees,
        "tasks": tasks,
        "completed": completed,
        "productivity": productivity
    }

@app.get("/search-employee")
def search_employee(name:str):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT
                id,
                full_name,
                email
                FROM users
                WHERE full_name ILIKE :name
            """),
            {
                "name":f"%{name}%"
            }
        )

        employees = [
            dict(row._mapping)
            for row in result
        ]

    return employees

@app.get("/employee-details")
def employee_details(user_id:int):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT
                id,
                full_name,
                email,
                role
                FROM users
                WHERE id=:id
            """),
            {"id":user_id}
        )

        employee = result.fetchone()

    if not employee:
        return {"success":False}

    return dict(employee._mapping)

@app.post("/approve-task")
def approve_task(task_id:int, _staff: dict = Depends(require_roles("admin", "manager"))):

    with engine.connect() as conn:

        # Fetch task details before updating so we can log to history
        task = conn.execute(
            text("""
                SELECT *
                FROM tasks
                WHERE id=:id
            """),
            {"id":task_id}
        ).fetchone()

        conn.execute(
            text("""
                UPDATE tasks
                SET status='Completed'
                WHERE id=:id
            """),
            {"id":task_id}
        )

        # Insert into task_history so the employee's History page is populated
        if task:
            conn.execute(
                text("""
                    INSERT INTO task_history
                    (task_id, employee_id, task_title, submitted_at)
                    VALUES
                    (:task_id, :employee_id, :task_title, :submitted_at)
                """),
                {
                    "task_id":      task.id,
                    "employee_id":  task.assigned_to,
                    "task_title":   task.title,
                    "submitted_at": datetime.now(ZoneInfo("Asia/Kolkata")).isoformat()
                }
            )

        conn.commit()

    return {"success":True}


@app.get("/review-tasks")
def review_tasks(manager_id: int):
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT
                    tasks.*,
                    users.full_name AS employee_name
                FROM tasks
                JOIN users ON users.id = tasks.assigned_to
                WHERE tasks.assigned_to IN (
                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id = :manager_id
                )
                AND LOWER(tasks.status) = 'pending review'
                ORDER BY tasks.id DESC
            """),
            {"manager_id": manager_id}
        ).fetchall()
    return [dict(row._mapping) for row in rows]


@app.post("/team/remove")
def remove_team_member(manager_id:int,
                       employee_id:int,
                       _staff: dict = Depends(require_roles("admin", "manager"))):

    print("manager =", manager_id)
    print("employee =", employee_id)

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                DELETE
                FROM team_members
                WHERE manager_id=:manager_id
                AND employee_id=:employee_id
            """),
            {
                "manager_id":manager_id,
                "employee_id":employee_id
            }
        )

        conn.commit()

    print("rows deleted =", result.rowcount)

    return {"success":True}


# ── TEAM REQUEST ENDPOINTS ───────────────────────────────────────

@app.get("/team-requests")
def get_team_requests(employee_id: int):
    """Employee sees all pending join requests sent to them."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT
                    tr.id,
                    tr.manager_id,
                    tr.status,
                    tr.created_at,
                    u.full_name AS manager_name,
                    u.email     AS manager_email
                FROM team_requests tr
                JOIN users u ON u.id = tr.manager_id
                WHERE tr.employee_id = :employee_id
                AND   tr.status      = 'pending'
                ORDER BY tr.created_at DESC
            """),
            {"employee_id": employee_id}
        ).fetchall()
    return [dict(row._mapping) for row in rows]


@app.get("/team-requests/sent")
def get_sent_requests(manager_id: int):
    """Manager sees all requests they have sent and their current status."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT
                    tr.id,
                    tr.employee_id,
                    tr.status,
                    tr.created_at,
                    u.full_name AS employee_name,
                    u.email     AS employee_email
                FROM team_requests tr
                JOIN users u ON u.id = tr.employee_id
                WHERE tr.manager_id = :manager_id
                ORDER BY tr.created_at DESC
            """),
            {"manager_id": manager_id}
        ).fetchall()
    return [dict(row._mapping) for row in rows]


@app.post("/team-request/respond")
def respond_team_request(request_id: int, action: str):
    """Employee approves or rejects a pending team join request."""
    if action not in ("approved", "rejected"):
        return {"success": False, "message": "Invalid action. Use 'approved' or 'rejected'."}

    with engine.begin() as conn:
        request = conn.execute(
            text("SELECT * FROM team_requests WHERE id = :id"),
            {"id": request_id}
        ).fetchone()

        if not request:
            return {"success": False, "message": "Request not found."}

        if request.status != "pending":
            return {"success": False, "message": "Request has already been responded to."}

        conn.execute(
            text("UPDATE team_requests SET status = :status WHERE id = :id"),
            {"status": action, "id": request_id}
        )

        if action == "approved":
            # Check not already on team (safety guard)
            existing = conn.execute(
                text("""
                    SELECT id FROM team_members
                    WHERE manager_id  = :manager_id
                    AND   employee_id = :employee_id
                """),
                {"manager_id": request.manager_id, "employee_id": request.employee_id}
            ).fetchone()

            if not existing:
                conn.execute(
                    text("""
                        INSERT INTO team_members (manager_id, employee_id)
                        VALUES (:manager_id, :employee_id)
                    """),
                    {"manager_id": request.manager_id, "employee_id": request.employee_id}
                )

    return {"success": True}


@app.get("/team")
def get_team(manager_id:int):

    with engine.connect() as conn:

        data = conn.execute(
            text("""
                SELECT
                    users.id,
                    users.full_name,
                    users.email,

                    COUNT(
                        CASE
                        WHEN LOWER(tasks.status) != 'completed'
                        THEN tasks.id
                        END
                    ) AS task_count

                FROM team_members

                JOIN users
                ON team_members.employee_id = users.id

                LEFT JOIN tasks
                ON tasks.assigned_to = users.id

                WHERE team_members.manager_id = :manager_id

                GROUP BY
                    users.id,
                    users.full_name,
                    users.email

                ORDER BY users.full_name
            """),
            {
                "manager_id":manager_id
            }
        ).fetchall()

    result = []

    for row in data:

        result.append({

            "id":row.id,

            "full_name":row.full_name,

            "email":row.email,

            "task_count":row.task_count

        })

    return result

@app.get("/manager-dashboard-stats")
def manager_dashboard_stats(manager_id:int):

    with engine.connect() as conn:

        # Team members
        team_members = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM team_members
                WHERE manager_id=:manager_id
            """),
            {"manager_id":manager_id}
        ).scalar()

        # Total tasks assigned to team
        total_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to IN(
                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id
                )
            """),
            {"manager_id":manager_id}
        ).scalar()

        # Completed tasks
        completed_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to IN(
                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id
                )
                AND LOWER(status)='completed'
            """),
            {"manager_id":manager_id}
        ).scalar()

        # Pending tasks (not completed)
        pending_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to IN(
                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id
                )
                AND LOWER(status)!='completed'
            """),
            {"manager_id":manager_id}
        ).scalar()

        # Tasks pending review
        review_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to IN(
                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id
                )
                AND LOWER(status)='pending review'
            """),
            {"manager_id":manager_id}
        ).scalar()

    # Productivity = completed / total * 100
    productivity = 0
    if total_tasks and total_tasks > 0:
        productivity = round(completed_tasks * 100 / total_tasks)

    return {
        "team_members":   team_members   or 0,
        "total_tasks":    total_tasks    or 0,
        "completed_tasks": completed_tasks or 0,
        "pending_tasks":  pending_tasks  or 0,
        "review_tasks":   review_tasks   or 0,
        "productivity":   productivity
    }

@app.get("/manager-tasks")
def manager_tasks(manager_id:int):

    with engine.connect() as conn:

        rows = conn.execute(
            text("""
                SELECT
                    tasks.id,
                    tasks.title,
                    users.full_name AS employee_name,
                    tasks.assigned_to,
                    tasks.priority,
                    tasks.status,
                    tasks.due_date

                FROM tasks

                JOIN users
                ON tasks.assigned_to = users.id

                WHERE tasks.assigned_to IN (

                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id

                )

                ORDER BY tasks.id DESC
            """),
            {"manager_id": manager_id}
        ).fetchall()

    return [dict(row._mapping) for row in rows]

@app.post("/team/add")
def add_team_member(manager_id: int, employee_id: int, _staff: dict = Depends(require_roles("admin", "manager"))):
    with engine.connect() as conn:
        existing = conn.execute(
            text("""
                SELECT id FROM team_requests
                WHERE manager_id  = :manager_id
                AND   employee_id = :employee_id
                AND   status      = 'pending'
            """),
            {"manager_id": manager_id, "employee_id": employee_id}
        ).fetchone()
        if existing:
            return {"success": False, "message": "A request is already pending for this employee."}
        on_team = conn.execute(
            text("""
                SELECT id FROM team_members
                WHERE manager_id  = :manager_id
                AND   employee_id = :employee_id
            """),
            {"manager_id": manager_id, "employee_id": employee_id}
        ).fetchone()
        if on_team:
            return {"success": False, "message": "This employee is already on your team."}
        conn.execute(
            text("""
                INSERT INTO team_requests (manager_id, employee_id, status)
                VALUES (:manager_id, :employee_id, 'pending')
            """),
            {"manager_id": manager_id, "employee_id": employee_id}
        )
        conn.commit()
    return {"success": True}

@app.get("/manager-report")
def manager_report(manager_id:int):

    with engine.connect() as conn:

        # Team members
        team_members = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM team_members
                WHERE manager_id=:manager_id
            """),
            {"manager_id":manager_id}
        ).scalar()

        # Assigned tasks
        assigned_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to IN(

                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id

                )
            """),
            {"manager_id":manager_id}
        ).scalar()

        # Completed tasks
        completed_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to IN(

                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id

                )

                AND LOWER(status)='completed'
            """),
            {"manager_id":manager_id}
        ).scalar()

        # Pending tasks
        pending_tasks = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM tasks
                WHERE assigned_to IN(

                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id

                )

                AND LOWER(status)!='completed'
            """),
            {"manager_id":manager_id}
        ).scalar()

        # Completion rate
        if assigned_tasks == 0:
            completion_rate = 0
        else:
            completion_rate = round(
                completed_tasks * 100 / assigned_tasks
            )

        # Top performer
        top_performer_row = conn.execute(
            text("""
                SELECT
                    users.full_name,
                    COUNT(tasks.id) AS completed_count

                FROM users

                JOIN tasks
                ON users.id = tasks.assigned_to

                WHERE users.id IN(

                    SELECT employee_id
                    FROM team_members
                    WHERE manager_id=:manager_id

                )

                AND LOWER(tasks.status)='completed'

                GROUP BY users.full_name

                ORDER BY completed_count DESC

                LIMIT 1
            """),
            {"manager_id":manager_id}
        ).fetchone()

        top_performer = (
            top_performer_row.full_name
            if top_performer_row
            else "N/A"
        )

    return {

        "team_members": team_members or 0,

        "assigned_tasks": assigned_tasks or 0,

        "completed_tasks": completed_tasks or 0,

        "pending_tasks": pending_tasks or 0,

        "completion_rate": completion_rate,

        "top_performer": top_performer

    }
@app.post("/delete-user")
def delete_user(user_id: int, _admin: dict = Depends(require_roles("admin"))):

    with engine.connect() as conn:

        # Delete tasks assigned to employee
        conn.execute(
            text("""
                DELETE FROM tasks
                WHERE assigned_to=:user_id
            """),
            {"user_id": user_id}
        )

        # Remove employee from teams
        conn.execute(
            text("""
                DELETE FROM team_members
                WHERE employee_id=:user_id
            """),
            {"user_id": user_id}
        )

        # Delete the user
        conn.execute(
            text("""
                DELETE FROM users
                WHERE id=:user_id
            """),
            {"user_id": user_id}
        )

        conn.commit()

    return {
        "success": True,
        "message": "User deleted successfully"
    }
@app.get("/managers")
def get_managers():

    with engine.connect() as conn:

        rows = conn.execute(
            text("""
                SELECT
                    id,
                    full_name,
                    email,
                    role
                FROM users
                WHERE LOWER(role)='manager'
                ORDER BY id
            """)
        ).fetchall()

    return [dict(row._mapping) for row in rows]


# ── UPDATE PROFILE ───────────────────────────────────────────────
@app.post("/update-profile")
def update_profile(user_id: int, full_name: str, email: str):
    with engine.connect() as conn:
 
        # Check email not taken by someone else
        existing = conn.execute(
            text("SELECT id FROM users WHERE email=:email AND id != :user_id"),
            {"email": email, "user_id": user_id}
        ).fetchone()
 
        if existing:
            return {"success": False, "message": "Email already in use by another account"}
 
        conn.execute(
            text("""
                UPDATE users
                SET full_name = :full_name, email = :email
                WHERE id = :user_id
            """),
            {"full_name": full_name, "email": email, "user_id": user_id}
        )
        conn.commit()
 
    return {"success": True, "message": "Profile updated"}
 
 
# ── CHANGE PASSWORD ──────────────────────────────────────────────
class ChangePasswordRequest(BaseModel):
    user_id: int
    current_password: str
    new_password: str


@app.post("/change-password")
def change_password(data: ChangePasswordRequest):
    with engine.connect() as conn:
 
        # Verify current password
        user = conn.execute(
            text("SELECT id, password FROM users WHERE id=:user_id"),
            {"user_id": data.user_id}
        ).fetchone()
 
        if not user or not verify_password(data.current_password, user.password):
            return {"success": False, "message": "Current password is incorrect"}
 
        if len(data.new_password) < 6:
            return {"success": False, "message": "New password must be at least 6 characters"}
 
        conn.execute(
            text("UPDATE users SET password=:password WHERE id=:user_id"),
            {"password": hash_password(data.new_password), "user_id": data.user_id}
        )
        conn.commit()
 
    return {"success": True, "message": "Password changed successfully"}
 
 
# ── GET ALL USERS WITH ROLES (for role management) ───────────────
@app.get("/all-users")
def get_all_users():
    with engine.connect() as conn:
        result = conn.execute(
            text("""
                SELECT id, full_name, email, role
                FROM users
                ORDER BY role, full_name
            """)
        )
        users = [dict(row._mapping) for row in result]
    return users
 
 
# ── UPDATE USER ROLE ─────────────────────────────────────────────
@app.post("/update-role")
def update_role(user_id: int, new_role: str, _admin: dict = Depends(require_roles("admin"))):
    valid_roles = ["employee", "manager", "admin"]
    if new_role not in valid_roles:
        return {"success": False, "message": "Invalid role"}
 
    with engine.connect() as conn:
        conn.execute(
            text("UPDATE users SET role=:role WHERE id=:user_id"),
            {"role": new_role, "user_id": user_id}
        )
        conn.commit()
 
    return {"success": True}



# ==========================================================
# SALES DASHBOARD ROUTES
# ==========================================================

from typing import Optional


# ----------------------------------------------------------
# TEAM MEMBER IDS
# ----------------------------------------------------------

def _team_member_ids(manager_id: int, conn):

    rows = conn.execute(
        text("""
            SELECT employee_id
            FROM team_members
            WHERE manager_id=:manager_id
        """),
        {"manager_id": manager_id}
    ).fetchall()

    return [row.employee_id for row in rows]


# ----------------------------------------------------------
# CURRENT WEEK
# ----------------------------------------------------------

@app.get("/sales/current-week")
def sales_current_week():

    return {

        "week": current_week()

    }


# ----------------------------------------------------------
# ADD PROJECTION
# ----------------------------------------------------------

from typing import Optional

@app.post("/sales/projection")
def post_projection(
    user_id: int,
    week: int,
    customer: Optional[str] = "",
    product: Optional[str] = "",
    projected: Optional[int] = 0,
    price: Optional[float] = 0
):

    if week < current_week():

        return {

            "success": False,
            "message": "Projections can only be added for current week."

        }

    try:

        return add_projection(

            user_id=user_id,
            week=week,
            customer=customer,
            product=product,
            projected=projected,
            price=price

        )

    except Exception as e:

        return {

            "success": False,
            "message": str(e)

        }


# ----------------------------------------------------------
# GET SALES ROWS
# ----------------------------------------------------------

@app.get("/sales/rows")
def sales_rows(
    user_id: int,
    week: int,
    viewer_id: int,
    viewer_role: str
):
    if viewer_role == "admin":
        pass  # admin can view anyone
    elif viewer_role == "manager":
        # manager can view their own rows OR their team members'
        if viewer_id != user_id:
            with engine.connect() as conn:
                team = _team_member_ids(viewer_id, conn)
            if user_id not in team:
                return {"success": False, "message": "Not your team member"}
    else:
        # employee can only view their own rows
        if viewer_id != user_id:
            return {"success": False, "message": "Access denied"}

    return {"rows": get_sales(user_id, week)}


# ----------------------------------------------------------
# GET WEEKS
# ----------------------------------------------------------

@app.get("/sales/weeks")
def sales_weeks(
    user_id: int,
    viewer_id: int,
    viewer_role: str
):
    if viewer_role == "admin":
        pass  # admin can view anyone
    elif viewer_role == "manager":
        # manager can view own weeks OR team members'
        if viewer_id != user_id:
            with engine.connect() as conn:
                team = _team_member_ids(viewer_id, conn)
            if user_id not in team:
                return {"success": False}
    else:
        if viewer_id != user_id:
            return {"success": False}

    return {"weeks": get_all_weeks(user_id)}


# ----------------------------------------------------------
# UPDATE ACHIEVED
# ----------------------------------------------------------

@app.put("/sales/achieved")
def sales_achieved(
    user_id: int,
    week: int,
    customer: str,
    product: str,
    achieved: int
):

    if week < current_week():
        return {
            "success": False,
            "message": "Only current or future weeks are editable."
        }

    try:

        return update_achieved(

            user_id,
            week,
            customer,
            product,
            achieved

        )

    except Exception as e:

        return {

            "success": False,
            "message": str(e)

        }


# ----------------------------------------------------------
# TEAM SALES
# ----------------------------------------------------------

@app.get("/sales/team")
def sales_team(
    manager_id: int,
    week: int,
    viewer_role: str
):

    with engine.connect() as conn:

        if viewer_role == "admin":

            rows = conn.execute(
                text("""
                    SELECT id
                    FROM users
                    WHERE role='employee'
                """)
            ).fetchall()

            employee_ids = [row.id for row in rows]

        else:

            employee_ids = _team_member_ids(
                manager_id,
                conn
            )

    data = get_sales_for_employees(
        employee_ids,
        week
    )

    names = {}

    if employee_ids:

        placeholders = ",".join(
            str(i)
            for i in employee_ids
        )

        with engine.connect() as conn:

            rows = conn.execute(
                text(f"""
                    SELECT
                    id,
                    full_name

                    FROM users

                    WHERE id IN ({placeholders})
                """)
            ).fetchall()

        names = {

            row.id: row.full_name

            for row in rows

        }

    result = []

    for emp_id in employee_ids:

        result.append({

            "id": emp_id,
            "name": names.get(
                emp_id,
                f"Employee {emp_id}"
            ),
            "rows": data.get(
                emp_id,
                []
            )

        })

    return {

        "week": week,
        "employees": result

    }


# ----------------------------------------------------------
# ADMIN / MANAGER EDIT PROJECTION
# ----------------------------------------------------------

@app.put("/sales/projection/admin-edit")
def admin_edit_projection_route(
    viewer_role: str,
    user_id: int,
    week: int,
    customer: str,
    product: str,
    new_qty: int
):

    if viewer_role not in [

        "admin",
        "manager"

    ]:

        return {

            "success": False

        }

    try:

        return admin_update_projection(

            user_id,
            week,
            customer,
            product,
            new_qty

        )

    except Exception as e:

        return {

            "success": False,
            "message": str(e)

        }


# ----------------------------------------------------------
# CHANGE REQUEST
# ----------------------------------------------------------

@app.post("/sales/change-request")
def sales_change_request(
    employee_id: int,
    week: int,
    customer: str,
    product: str,
    old_qty: int,
    new_qty: int,
    reason: str
):

    return raise_change_request(

        employee_id,
        week,
        customer,
        product,
        old_qty,
        new_qty,
        reason

    )


# ----------------------------------------------------------
# LIST CHANGE REQUESTS
# ----------------------------------------------------------

@app.get("/sales/change-requests")
def list_change_requests(
    viewer_id: int,
    viewer_role: str,
    status: Optional[str] = None
):

    if viewer_role == "employee":

        requests = get_change_requests(

            employee_ids=[viewer_id],
            status=status

        )

    elif viewer_role == "manager":

        with engine.connect() as conn:

            team = _team_member_ids(
                viewer_id,
                conn
            )

        requests = get_change_requests(

            employee_ids=team,
            status=status

        )

    elif viewer_role == "admin":

        requests = get_change_requests(

            employee_ids=None,
            status=status

        )

    else:

        requests = []

    return {

        "requests": requests

    }


# ----------------------------------------------------------
# RESOLVE CHANGE REQUEST
# ----------------------------------------------------------

@app.post("/sales/change-request/resolve")
def resolve_request(
    request_id: str,
    action: str,
    manager_note: str = "",
    viewer_role: str = ""
):

    if viewer_role not in [

        "admin",
        "manager"

    ]:

        return {

            "success": False

        }

    try:

        return resolve_change_request(

            request_id,
            action,
            manager_note

        )

    except Exception as e:

        return {

            "success": False,
            "message": str(e)

        }

from fastapi.responses import FileResponse

from fastapi.responses import FileResponse

@app.get("/sales/download")
def download_sales():

    return FileResponse(
        FILE_NAME,
        filename="sales_tracker.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

from openpyxl import load_workbook

@app.get("/sales/debug")
def sales_debug():

    wb = load_workbook(FILE_NAME)
    output = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [row for row in ws.iter_rows(values_only=True)]
        output[sheet] = rows

    return output
@app.get("/sales/debug-path")
def debug_path():

    import os

    return {

        "cwd": os.getcwd(),

        "files": os.listdir()

    }
@app.get("/test123")
def test123():
    return {
        "ok": True
    }
import os

@app.get("/testfiles")
def testfiles():

    try:

        return {
            "cwd": os.getcwd(),
            "files": os.listdir(".")
        }

    except Exception as e:

        return {
            "error": str(e)
        }
@app.get("/sales/test-add")
def sales_test_add():

    try:

        return add_projection(
            user_id=11,
            week=current_week(),
            customer="XYZ Company",
            product="Switch",
            projected=25,
            price=800
        )

    except Exception as e:

        return {
            "success": False,
            "error": str(e)
        }


# ═══════════════════════════════════════════════════════
# TASK FILE ATTACHMENT
# ═══════════════════════════════════════════════════════
# Run once to add the column (idempotent):
#   ALTER TABLE tasks ADD COLUMN IF NOT EXISTS file_url TEXT;
#   ALTER TABLE tasks ADD COLUMN IF NOT EXISTS file_name TEXT;

import shutil, os as _os

# Files now live in the database (stored_files table) instead of a local/volume
# path, so they survive on any host -- no persistent disk required.
with engine.begin() as _conn:
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS stored_files (
            filename      TEXT PRIMARY KEY,
            content_type  TEXT,
            data          BYTEA NOT NULL,
            original_name TEXT,
            created_at    TIMESTAMP DEFAULT now()
        )
    """))

@app.post("/submit-task-with-file")
async def submit_task_with_file(
    task_id: int,
    file: UploadFile = FastAPIFile(None)
):
    """
    Employee calls this instead of /update-task-status when submitting
    for review with an optional file attachment.
    Saves the file into the database and stores the path in tasks.
    """
    file_url  = None
    file_name = None

    if file and file.filename:
        safe_name = f"{task_id}_{file.filename.replace(' ', '_')}"
        contents  = await file.read()
        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO stored_files (filename, content_type, data, original_name)
                    VALUES (:fn, :ct, :data, :orig)
                    ON CONFLICT (filename) DO UPDATE
                    SET content_type = EXCLUDED.content_type,
                        data = EXCLUDED.data,
                        original_name = EXCLUDED.original_name,
                        created_at = now()
                """),
                {"fn": safe_name, "ct": file.content_type, "data": contents, "orig": file.filename}
            )
        file_url  = f"/task-file/{safe_name}"
        file_name = file.filename

    with engine.connect() as conn:
        conn.execute(
            text("""
                UPDATE tasks
                SET status    = 'Pending Review',
                    file_url  = :file_url,
                    file_name = :file_name
                WHERE id = :task_id
            """),
            {"task_id": task_id, "file_url": file_url, "file_name": file_name}
        )
        conn.commit()

    return {"success": True, "file_url": file_url, "file_name": file_name}


from fastapi.responses import Response as _RawResponse

@app.get("/task-file/{filename}")
def serve_task_file(filename: str):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT content_type, data, original_name FROM stored_files WHERE filename=:fn"),
            {"fn": filename}
        ).mappings().first()
    if not row:
        return {"error": "File not found"}
    return _RawResponse(
        content=bytes(row["data"]),
        media_type=row["content_type"] or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{row["original_name"] or filename}"'}
    )


# ═══════════════════════════════════════════════════════
# TASK REMARKS (manager → employee corrections)
# ═══════════════════════════════════════════════════════
# Run once:
#   CREATE TABLE IF NOT EXISTS task_remarks (
#       id          SERIAL PRIMARY KEY,
#       task_id     INT NOT NULL,
#       manager_id  INT NOT NULL,
#       employee_id INT NOT NULL,
#       remark      TEXT NOT NULL,
#       status      TEXT DEFAULT 'pending',   -- pending | resolved
#       created_at  TIMESTAMP DEFAULT NOW()
#   );

@app.post("/send-remark")
def send_remark(
    task_id:     int,
    manager_id:  int,
    employee_id: int,
    remark:      str
):
    """
    Manager sends a correction remark on a task.
    Also resets the task status back to 'In Progress' so employee
    can fix and re-submit.
    """
    with engine.connect() as conn:

        conn.execute(
            text("""
                INSERT INTO task_remarks
                (task_id, manager_id, employee_id, remark, status)
                VALUES
                (:task_id, :manager_id, :employee_id, :remark, 'pending')
            """),
            {
                "task_id":     task_id,
                "manager_id":  manager_id,
                "employee_id": employee_id,
                "remark":      remark
            }
        )

        # Push task back to In Progress so employee can re-submit
        conn.execute(
            text("""
                UPDATE tasks
                SET status = 'In Progress'
                WHERE id = :task_id
            """),
            {"task_id": task_id}
        )

        conn.commit()

    return {"success": True}


@app.get("/task-remarks")
def get_task_remarks(employee_id: int):
    """
    Returns all pending remarks for an employee across all tasks.
    """
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT
                    tr.id,
                    tr.task_id,
                    tr.remark,
                    tr.status,
                    tr.created_at,
                    t.title  AS task_title,
                    u.full_name AS manager_name
                FROM task_remarks tr
                JOIN tasks t  ON tr.task_id    = t.id
                JOIN users u  ON tr.manager_id = u.id
                WHERE tr.employee_id = :employee_id
                ORDER BY tr.created_at DESC
            """),
            {"employee_id": employee_id}
        ).fetchall()

    return [dict(r._mapping) for r in rows]


@app.put("/resolve-remark")
def resolve_remark(remark_id: int):
    """Employee marks a remark as resolved after fixing."""
    with engine.connect() as conn:
        conn.execute(
            text("UPDATE task_remarks SET status='resolved' WHERE id=:id"),
            {"id": remark_id}
        )
        conn.commit()
    return {"success": True}


@app.get("/manager-remarks")
def get_manager_remarks(manager_id: int):
    """Returns all remarks sent by this manager, with current task status."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT
                    tr.id,
                    tr.task_id,
                    tr.remark,
                    tr.status      AS remark_status,
                    tr.created_at,
                    t.title        AS task_title,
                    t.status       AS task_status,
                    u.full_name    AS employee_name
                FROM task_remarks tr
                JOIN tasks t  ON tr.task_id     = t.id
                JOIN users u  ON tr.employee_id = u.id
                WHERE tr.manager_id = :manager_id
                ORDER BY tr.created_at DESC
            """),
            {"manager_id": manager_id}
        ).fetchall()

    return [dict(r._mapping) for r in rows]

@app.post("/meetings/create")
def create_meeting(
    title: str,
    description: str = "",
    meeting_date: str = "",
    start_slot: int = 0,
    end_slot: int = 0,
    organizer_id: int = 0,
    location: str = "",
    attendees: str = "[]"
):
    import json
    attendee_list = json.loads(attendees)

    with engine.begin() as conn:
        conflict = conn.execute(
            text("""
                SELECT id FROM meetings
                WHERE meeting_date = :meeting_date
                AND organizer_id = :organizer_id
                AND (start_slot < :new_end AND end_slot > :new_start)
            """),
            {"meeting_date": meeting_date, "organizer_id": organizer_id,
             "new_start": start_slot, "new_end": end_slot}
        ).fetchone()

        if conflict:
            return {"success": False, "message": "You already have a meeting at that time"}

        meeting_id = conn.execute(
            text("""
                INSERT INTO meetings
                    (title, description, organizer_id, meeting_date, start_slot, end_slot, location)
                VALUES
                    (:title, :description, :organizer_id, :meeting_date, :start_slot, :end_slot, :location)
                RETURNING id
            """),
            {"title": title, "description": description, "organizer_id": organizer_id,
             "meeting_date": meeting_date, "start_slot": start_slot,
             "end_slot": end_slot, "location": location}
        ).scalar()

        for uid in attendee_list:
            conn.execute(
                text("INSERT INTO meeting_attendees(meeting_id, user_id) VALUES(:mid, :uid)"),
                {"mid": meeting_id, "uid": uid}
            )

    return {"success": True, "meeting_id": meeting_id}



@app.get("/meetings/month")
def get_month_meetings(user_id: int, month: str):
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT DISTINCT m.id, m.title, m.description,
                    m.meeting_date, m.start_slot, m.end_slot,
                    m.location, m.organizer_id
                FROM meetings m
                LEFT JOIN meeting_attendees a ON m.id = a.meeting_id
                WHERE (m.organizer_id = :user_id OR a.user_id = :user_id)
                AND TO_CHAR(m.meeting_date, 'YYYY-MM') = :month
                ORDER BY m.meeting_date, m.start_slot
            """),
            {"user_id": user_id, "month": month}
        ).fetchall()
    return [
        {"id": r.id, "title": r.title, "description": r.description,
         "meeting_date": str(r.meeting_date), "date": str(r.meeting_date),
         "start_slot": r.start_slot, "end_slot": r.end_slot,
         "location": r.location, "organizer_id": r.organizer_id}
        for r in rows
    ]
    
@app.get("/meetings/day")
def get_day_meetings(date: str, user_id: int):
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
            SELECT DISTINCT m.id, m.title, m.description,
                m.meeting_date, m.start_slot, m.end_slot,
                m.location, m.organizer_id
            FROM meetings m
            LEFT JOIN meeting_attendees a ON m.id = a.meeting_id
            WHERE m.meeting_date = :date
            AND (m.organizer_id = :user_id OR a.user_id = :user_id)
            ORDER BY m.start_slot
            """),
            {"date": date, "user_id": user_id}
        ).fetchall()

    return [
        {
            "id":           row.id,
            "title":        row.title,
            "description":  row.description,
            "date":         str(row.meeting_date),
            "start_slot":   row.start_slot,
            "end_slot":     row.end_slot,
            "start_time":   slot_to_time(row.start_slot),
            "end_time":     slot_to_time(row.end_slot),
            "location":     row.location,
            "organizer_id": row.organizer_id,
        }
        for row in rows
    ]

@app.get("/meetings/user-availability")
def user_availability(user_id:int, date:str):

    availability = [True]*48

    with engine.connect() as conn:

        rows = conn.execute(
            text("""
            SELECT
                start_slot,
                end_slot
            FROM meetings
            WHERE
                meeting_date=:date
            AND
                organizer_id=:user_id
            """),
            {
                "date": date,
                "user_id": user_id
            }
        ).fetchall()

    for row in rows:

        for slot in range(row.start_slot,row.end_slot):

            availability[slot] = False

    return {
        "slots": availability
    }

@app.get("/meetings/week")
def get_week_meetings(user_id: int, week_start: str):

    with engine.connect() as conn:

        rows = conn.execute(
            text("""
            SELECT
                m.id,
                m.title,
                m.description,
                m.meeting_date,
                m.start_slot,
                m.end_slot,
                m.location,
                m.organizer_id
            FROM meetings m

            LEFT JOIN meeting_attendees a
            ON m.id = a.meeting_id

            WHERE
            (
                m.organizer_id=:user_id
                OR
                a.user_id=:user_id
            )

            AND

            m.meeting_date BETWEEN
            :week_start::date
            AND
            (:week_start::date + interval '6 days')

            ORDER BY
            m.meeting_date,
            m.start_slot
            """),
            {
                "user_id": user_id,
                "week_start": week_start
            }
        ).fetchall()

    output = []

    for row in rows:

        output.append({

            "id": row.id,

            "title": row.title,

            "description": row.description,

            "date": str(row.meeting_date),

            "start_slot": row.start_slot,

            "end_slot": row.end_slot,

            "start_time": slot_to_time(row.start_slot),

            "end_time": slot_to_time(row.end_slot),

            "location": row.location,

            "organizer_id": row.organizer_id

        })

    return output

@app.get("/meetings/{meeting_id}/attendees")
def get_meeting_attendees(meeting_id: int):
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT u.id AS user_id, u.full_name
                FROM meeting_attendees ma
                JOIN users u ON u.id = ma.user_id
                WHERE ma.meeting_id = :meeting_id
            """),
            {"meeting_id": meeting_id}
        ).fetchall()
    return [{"user_id": r.user_id, "full_name": r.full_name} for r in rows]

@app.get("/meeting/{meeting_id}")
def get_meeting_details(meeting_id: int):

    with engine.connect() as conn:

        row = conn.execute(
            text("""
            SELECT *
            FROM meetings
            WHERE id=:meeting_id
            """),
            {
                "meeting_id": meeting_id
            }
        ).fetchone()

        attendees = conn.execute(
            text("""
            SELECT
                u.id,
                u.full_name,
                ma.response_status
            FROM meeting_attendees ma

            JOIN users u
            ON u.id=ma.user_id

            WHERE ma.meeting_id=:meeting_id
            """),
            {
                "meeting_id": meeting_id
            }
        ).fetchall()

    if not row:

        return {"success": False}

    return {

        "success": True,

        "meeting": {

            "id": row.id,

            "title": row.title,

            "description": row.description,

            "date": str(row.meeting_date),

            "start_slot": row.start_slot,

            "end_slot": row.end_slot,

            "start_time": slot_to_time(row.start_slot),

            "end_time": slot_to_time(row.end_slot),

            "location": row.location,

            "organizer_id": row.organizer_id

        },

        "attendees": [

            {

                "id": a.id,

                "name": a.full_name,

                "status": a.response_status

            }

            for a in attendees

        ]

    }

class MeetingResponseRequest(BaseModel):

    meeting_id: int

    user_id: int

    status: str

@app.post("/meeting/respond")
def respond_to_meeting(req: MeetingResponseRequest):

    with engine.begin() as conn:

        conn.execute(

            text("""
            UPDATE meeting_attendees

            SET response_status=:status

            WHERE

            meeting_id=:meeting_id

            AND

            user_id=:user_id
            """),

            req.model_dump()

        )

    return {

        "success": True

    }
@app.get("/calendar-tasks")
def get_calendar_tasks(user_id: int, month: Optional[str] = None, date: Optional[str] = None):
    """
    Supports two modes:
      ?user_id=X&month=YYYY-MM   → all tasks for a month (month view)
      ?user_id=X&date=YYYY-MM-DD → all tasks for a single day (day view)
    """
    try:
        with engine.connect() as conn:

            if date:
                rows = conn.execute(
                    text("""
                    SELECT
                        id, title, description, task_date,
                        start_slot, end_slot, color, status
                    FROM calendar_tasks
                    WHERE user_id = :user_id
                    AND task_date = :date
                    ORDER BY start_slot
                    """),
                    {"user_id": user_id, "date": date}
                ).fetchall()
            else:
                rows = conn.execute(
                    text("""
                    SELECT
                        id, title, description, task_date,
                        start_slot, end_slot, color, status
                    FROM calendar_tasks
                    WHERE user_id = :user_id
                    AND TO_CHAR(task_date, 'YYYY-MM') = :month
                    ORDER BY task_date, start_slot
                    """),
                    {"user_id": user_id, "month": month or ""}
                ).fetchall()

        return [
            {
                "id":          row.id,
                "title":       row.title,
                "description": row.description,
                "task_date":   str(row.task_date),
                "start_slot":  row.start_slot,
                "end_slot":    row.end_slot,
                "color":       row.color,
                "status":      row.status,
            }
            for row in rows
        ]
    except Exception as e:
        return []

# ═══════════════════════════════════════════════════════
# CALENDAR TASKS — CREATE / DELETE
# ═══════════════════════════════════════════════════════
# Run once in your DB if the table doesn't exist yet:
#   CREATE TABLE IF NOT EXISTS calendar_tasks (
#       id          SERIAL PRIMARY KEY,
#       user_id     INT NOT NULL,
#       title       TEXT NOT NULL,
#       description TEXT DEFAULT '',
#       task_date   DATE NOT NULL,
#       start_slot  INT NOT NULL,
#       end_slot    INT NOT NULL,
#       color       TEXT DEFAULT '#7c3aed',
#       status      TEXT DEFAULT 'pending',
#       created_at  TIMESTAMP DEFAULT NOW()
#   );

@app.post("/calendar-tasks/create")
def create_calendar_task(
    user_id: int,
    title: str,
    description: str = "",
    task_date: str = "",
    start_slot: int = 0,
    end_slot: int = 0,
    color: str = "#7c3aed",
    sync_gcal: bool = False
):
    try:
        with engine.begin() as conn:
            task_id = conn.execute(
                text("""
                    INSERT INTO calendar_tasks
                        (user_id, title, description, task_date, start_slot, end_slot, color)
                    VALUES
                        (:user_id, :title, :description, :task_date, :start_slot, :end_slot, :color)
                    RETURNING id
                """),
                {"user_id": user_id, "title": title, "description": description,
                 "task_date": task_date, "start_slot": start_slot,
                 "end_slot": end_slot, "color": color}
            ).scalar()
        return {"success": True, "task_id": task_id, "gcal_synced": False}
    except Exception as e:
        print(f"[ERROR] create_calendar_task failed: {e}", flush=True)
        return {"success": False, "message": str(e)}


@app.delete("/calendar-tasks/{task_id}")
def delete_calendar_task(task_id: int, user_id: int):
    with engine.begin() as conn:
        result = conn.execute(
            text("""
                DELETE FROM calendar_tasks
                WHERE id = :task_id AND user_id = :user_id
            """),
            {"task_id": task_id, "user_id": user_id}
        )
    if result.rowcount == 0:
        return {"success": False, "message": "Task not found or not yours"}
    return {"success": True}



@app.get("/calendar/combined")
def calendar_combined(user_id: int, date: str):

    with engine.connect() as conn:

        meetings = conn.execute(
            text("""
            SELECT DISTINCT m.title, m.start_slot, m.end_slot, 'meeting' as type
            FROM meetings m
            LEFT JOIN meeting_attendees a ON m.id = a.meeting_id
            WHERE m.meeting_date = :date
            AND (m.organizer_id = :user_id OR a.user_id = :user_id)
            """),
            {"date": date, "user_id": user_id}
        ).fetchall()

        tasks = conn.execute(
            text("""
            SELECT
                title,
                start_slot,
                end_slot,
                'task' as type
            FROM calendar_tasks
            WHERE
            user_id=:user_id
            AND task_date=:date
            """),
            {
                "user_id": user_id,
                "date": date
            }
        ).fetchall()

    output = []

    for row in meetings:
        output.append(dict(row._mapping))

    for row in tasks:
        output.append(dict(row._mapping))

    return output

@app.get("/users")
def get_all_users():
    with engine.connect() as conn:
        result = conn.execute(
            text("SELECT id, full_name, email, role FROM users ORDER BY full_name")
        )
        return [dict(row._mapping) for row in result]


RESEND_API_KEY = os.getenv("RESEND_API_KEY")


def send_email(to_email, subject, html):

    response = requests.post(
        "https://api.resend.com/emails",
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json"
        },
        json={
            "from": "Panache <onboarding@resend.dev>",
            "to": [to_email],
            "subject": subject,
            "html": html
        }
    )

    return response.json()


@app.post("/test-email")
def test_email():

    result = send_email(
        "krishpdgl@gmail.com",
        "Panache Test",
        """
        <h2>Email Working ✅</h2>
        <p>This email was sent using Resend.</p>
        """
    )

    return result


# ── WEB PUSH NOTIFICATIONS ────────────────────────────────
# VAPID key pair identifies this server to push services (Chrome/FCM etc).
# Generated once — the private key must stay secret. It's fine hardcoded
# here in the same way DATABASE_URL is, but can be moved to an env var
# (VAPID_PRIVATE_KEY) later without changing any other code.
VAPID_PRIVATE_KEY = os.getenv("VAPID_PRIVATE_KEY", """-----BEGIN PRIVATE KEY-----
MIGHAgEAMBMGByqGSM49AgEGCCqGSM49AwEHBG0wawIBAQQgVtRyMHbi7bvYntxV
4ABOS8En2ntbgGqgGVNR6fA9/+uhRANCAARJyJjPkQmN9eWwto5jq6Gefk2zVj9E
7Uqp77Bf5kz/Mms6h8S2ukGr1p4/oZ5nR1/DimNWSTyy2QMrlTYGBU50
-----END PRIVATE KEY-----
""")
VAPID_PUBLIC_KEY = os.getenv(
    "VAPID_PUBLIC_KEY",
    "BEnImM-RCY315bC2jmOroZ5-TbNWP0TtSqnvsF_mTP8yazqHxLa6QavWnj-hnmdHX8OKY1ZJPLLZAyuVNgYFTnQ"
)
VAPID_CLAIMS = {"sub": "mailto:admin@panache-wms.example"}


@app.get("/push/vapid-public-key")
def get_vapid_public_key():
    return {"publicKey": VAPID_PUBLIC_KEY}


class PushSubscriptionIn(BaseModel):
    user_id: int
    endpoint: str
    p256dh: str
    auth: str


@app.post("/push/subscribe")
def push_subscribe(data: PushSubscriptionIn):
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO push_subscriptions (user_id, endpoint, p256dh, auth)
                VALUES (:uid, :endpoint, :p256dh, :auth)
                ON CONFLICT (endpoint) DO UPDATE SET
                    user_id = EXCLUDED.user_id,
                    p256dh  = EXCLUDED.p256dh,
                    auth    = EXCLUDED.auth
            """),
            {"uid": data.user_id, "endpoint": data.endpoint, "p256dh": data.p256dh, "auth": data.auth}
        )
    return {"success": True}


class PushUnsubscribeIn(BaseModel):
    endpoint: str


@app.post("/push/unsubscribe")
def push_unsubscribe(data: PushUnsubscribeIn):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM push_subscriptions WHERE endpoint=:endpoint"), {"endpoint": data.endpoint})
    return {"success": True}


def send_push(user_id, title, body, url=None):
    """
    Sends a web push notification to every device this user is subscribed on.
    Never raises — a failed/expired subscription is silently cleaned up, and
    any other error is logged so it can never break the calling endpoint
    (task creation, payroll lock, leave approval, etc).
    """
    try:
        with engine.begin() as conn:
            subs = conn.execute(
                text("SELECT id, endpoint, p256dh, auth FROM push_subscriptions WHERE user_id=:uid"),
                {"uid": user_id}
            ).mappings().all()

            payload = json.dumps({"title": title, "body": body, "url": url or "index.html"})

            for sub in subs:
                try:
                    webpush(
                        subscription_info={
                            "endpoint": sub["endpoint"],
                            "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}
                        },
                        data=payload,
                        vapid_private_key=VAPID_PRIVATE_KEY,
                        vapid_claims=dict(VAPID_CLAIMS)
                    )
                except WebPushException as e:
                    status = getattr(e.response, "status_code", None)
                    if status in (404, 410):
                        # Subscription is dead (user uninstalled / cleared data) — remove it.
                        conn.execute(text("DELETE FROM push_subscriptions WHERE id=:id"), {"id": sub["id"]})
                    else:
                        print(f"[push warning] user {user_id}: {e}")
    except Exception as e:
        print(f"[send_push warning] user {user_id}: {e}")

from sqlalchemy import text

def send_sales_projection_reminders():

    with engine.connect() as conn:

        result = conn.execute(text("""
            SELECT full_name, email
            FROM users
            WHERE email IS NOT NULL
              AND email <> ''
        """))

        employees = result.fetchall()

    sent = 0
    failed = 0

    for emp in employees:

        try:

            full_name = emp.full_name
            email = emp.email

            response = send_email(
                email,
                "Sales Projection Reminder",
                f"""
                <h2>Weekly Sales Projection Reminder</h2>

                <p>Hi {full_name},</p>

                <p>Please fill your sales projections for this week.</p>

                <p>Login to Panache WMS and update your projections.</p>

                <p>Regards,<br>Panache WMS</p>
                """
            )

            if response.get("id"):
                sent += 1
            else:
                failed += 1
                print(f"Failed email: {email} -> {response}")

        except Exception as e:

            failed += 1

            print(f"Failed email: {email} -> {e}")

            continue

    return {
        "success": True,
        "sent": sent,
        "failed": failed,
        "total": len(employees)
    }

@app.post("/send-sales-reminders")
def send_sales_reminders():

    try:
        return send_sales_projection_reminders()

    except Exception as e:

        return {
            "success": False,
            "error": str(e)
        }
from sqlalchemy import text

def _resolve_approver(conn, user_id: int) -> Optional[int]:
    """
    Figure out who should approve a leave request for this user:
      - employee -> their manager (from team_members)
      - manager  -> an admin (managers' leaves go straight to admin)
      - admin    -> no approver (auto, nothing to route)
      - employee with no manager assigned -> fall back to an admin

    Defensive by design: if anything here goes wrong (e.g. an
    unexpected schema), we log it and return None rather than letting
    leave submission fail outright.
    """
    try:
        user_row = conn.execute(
            text("SELECT role FROM users WHERE id = :uid"),
            {"uid": user_id}
        ).mappings().first()

        role = (user_row["role"] if user_row else "") or ""
        role = role.lower()

        if role == "admin":
            return None

        if role == "manager":
            admin_row = conn.execute(
                text("SELECT id FROM users WHERE role = 'admin' ORDER BY id LIMIT 1")
            ).first()
            return admin_row[0] if admin_row else None

        # default: treat as employee -> look up their manager.
        # Primary source of truth is team_members, but as a safety net
        # also check for an approved team_requests row in case that
        # acceptance never got mirrored into team_members for some reason.
        manager_row = conn.execute(
            text("""
                SELECT manager_id
                FROM team_members
                WHERE employee_id = :uid
                ORDER BY id DESC
                LIMIT 1
            """),
            {"uid": user_id}
        ).first()

        if manager_row and manager_row[0]:
            return manager_row[0]

        manager_row = conn.execute(
            text("""
                SELECT manager_id
                FROM team_requests
                WHERE employee_id = :uid
                  AND status = 'approved'
                ORDER BY created_at DESC
                LIMIT 1
            """),
            {"uid": user_id}
        ).first()

        if manager_row and manager_row[0]:
            return manager_row[0]

        # truly no manager on record -> escalate straight to admin
        admin_row = conn.execute(
            text("SELECT id FROM users WHERE role = 'admin' ORDER BY id LIMIT 1")
        ).first()
        return admin_row[0] if admin_row else None
    except Exception as e:
        print(f"[_resolve_approver warning] could not resolve approver for user {user_id}: {e}")
        return None


# ── LEAVE TYPES & BALANCES ────────────────────────────────
@app.get("/leave-types")
def get_leave_types():
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT * FROM leave_types ORDER BY id")).mappings().all()
    return rows


class LeaveTypeIn(BaseModel):
    name: str
    annual_quota: Optional[float] = None
    carry_forward: bool = False
    is_balance_tracked: bool = True


@app.post("/leave-types")
def upsert_leave_type(data: LeaveTypeIn, _admin: dict = Depends(require_roles("admin"))):
    """Admin: create a new leave type, or update one if the name already exists."""
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO leave_types (name, annual_quota, carry_forward, is_balance_tracked)
                VALUES (:name, :quota, :cf, :bt)
                ON CONFLICT (name) DO UPDATE SET
                    annual_quota       = EXCLUDED.annual_quota,
                    carry_forward      = EXCLUDED.carry_forward,
                    is_balance_tracked = EXCLUDED.is_balance_tracked
            """),
            {"name": data.name, "quota": data.annual_quota, "cf": data.carry_forward, "bt": data.is_balance_tracked}
        )
    return {"success": True}


def _get_or_create_balance(conn, user_id, leave_type_id, year):
    """
    Lazily ensures a leave_balances row exists for this user/type/year,
    allocating from that leave type's annual_quota the first time it's
    touched. Must be called with an already-open (engine.begin()) conn.
    """
    row = conn.execute(
        text("SELECT allocated, used FROM leave_balances WHERE user_id=:uid AND leave_type_id=:ltid AND year=:yr"),
        {"uid": user_id, "ltid": leave_type_id, "yr": year}
    ).mappings().first()
    if row:
        return row

    lt = conn.execute(text("SELECT annual_quota FROM leave_types WHERE id=:id"), {"id": leave_type_id}).mappings().first()
    allocated = lt["annual_quota"] if lt and lt["annual_quota"] is not None else 0

    conn.execute(
        text("""
            INSERT INTO leave_balances (user_id, leave_type_id, year, allocated, used)
            VALUES (:uid, :ltid, :yr, :alloc, 0)
        """),
        {"uid": user_id, "ltid": leave_type_id, "yr": year, "alloc": allocated}
    )
    return {"allocated": allocated, "used": 0}


@app.get("/leave-balance")
def get_leave_balance(user_id: int, year: Optional[int] = None):
    yr = year or datetime.utcnow().year
    with engine.begin() as conn:
        types = conn.execute(text("SELECT id, name, is_balance_tracked FROM leave_types ORDER BY id")).mappings().all()
        out = []
        for lt in types:
            if not lt["is_balance_tracked"]:
                out.append({
                    "leave_type_id": lt["id"], "name": lt["name"],
                    "allocated": None, "used": None, "remaining": None
                })
                continue
            bal = _get_or_create_balance(conn, user_id, lt["id"], yr)
            allocated = float(bal["allocated"])
            used = float(bal["used"])
            out.append({
                "leave_type_id": lt["id"], "name": lt["name"],
                "allocated": allocated, "used": used, "remaining": allocated - used
            })
    return out


@app.post("/leave-requests")
def create_leave_request(data: LeaveRequest):

    # Resolve the approver in its own connection/transaction first, so
    # that if this lookup ever fails for any reason, it can't poison the
    # transaction used for the actual insert below.
    approver_id = None
    try:
        with engine.connect() as lookup_conn:
            approver_id = _resolve_approver(lookup_conn, data.user_id)
    except Exception as e:
        print(f"[create_leave_request warning] approver lookup failed: {e}")
        approver_id = None

    try:
        start_d = date.fromisoformat(data.start_date)
        end_d = date.fromisoformat(data.end_date)
    except ValueError:
        return {"success": False, "message": "Invalid date format."}

    days = (end_d - start_d).days + 1
    if days <= 0:
        return {"success": False, "message": "End date must be on or after the start date."}

    with engine.begin() as conn:

        # Balance check — only for leave types that are actually balance-tracked.
        if data.leave_type_id:
            lt = conn.execute(
                text("SELECT is_balance_tracked FROM leave_types WHERE id=:id"), {"id": data.leave_type_id}
            ).mappings().first()
            if lt and lt["is_balance_tracked"]:
                bal = _get_or_create_balance(conn, data.user_id, data.leave_type_id, start_d.year)
                remaining = float(bal["allocated"]) - float(bal["used"])
                if days > remaining:
                    return {
                        "success": False,
                        "message": f"Insufficient balance: requested {days} day(s) but only {remaining:g} remaining."
                    }

        result = conn.execute(
            text("""
                INSERT INTO leave_requests
                (
                    user_id,
                    employee_name,
                    leave_type,
                    leave_type_id,
                    start_date,
                    end_date,
                    reason,
                    approver_id,
                    days
                )
                VALUES
                (
                    :user_id,
                    :employee_name,
                    :leave_type,
                    :leave_type_id,
                    :start_date,
                    :end_date,
                    :reason,
                    :approver_id,
                    :days
                )
                RETURNING id
            """),
            {
                "user_id": data.user_id,
                "employee_name": data.employee_name,
                "leave_type": data.leave_type,
                "leave_type_id": data.leave_type_id,
                "start_date": data.start_date,
                "end_date": data.end_date,
                "reason": data.reason,
                "approver_id": approver_id,
                "days": days
            }
        )

        leave_id = result.scalar()

    if approver_id:
        try:
            send_push(approver_id, "New Leave Request", f"{data.employee_name} requested {days:g} day(s) of {data.leave_type}.")
        except Exception as e:
            print(f"[push warning] {e}")

    return {
        "success": True,
        "leave_id": leave_id,
        "days": days
    }

@app.get("/leave-requests")
def get_leave_requests():

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT *
                FROM leave_requests
                ORDER BY created_at DESC
            """)
        )

        rows = result.mappings().all()

    return rows

@app.post("/leave-requests/{leave_id}/approve")
def approve_leave(leave_id: int, _staff: dict = Depends(require_roles("admin", "manager"))):

    with engine.begin() as conn:

        lr = conn.execute(text("SELECT * FROM leave_requests WHERE id=:id"), {"id": leave_id}).mappings().first()
        if not lr:
            return {"success": False, "message": "Leave request not found."}

        conn.execute(
            text("""
                UPDATE leave_requests
                SET
                    status = 'Approved',
                    approved_at = NOW()
                WHERE id = :leave_id
            """),
            {
                "leave_id": leave_id
            }
        )

        # Deduct from balance, only if this request maps to a balance-tracked leave type.
        if lr["leave_type_id"] and lr["days"]:
            lt = conn.execute(
                text("SELECT is_balance_tracked FROM leave_types WHERE id=:id"), {"id": lr["leave_type_id"]}
            ).mappings().first()
            if lt and lt["is_balance_tracked"]:
                start_val = lr["start_date"]
                yr = start_val.year if hasattr(start_val, "year") else date.fromisoformat(str(start_val)).year
                _get_or_create_balance(conn, lr["user_id"], lr["leave_type_id"], yr)
                conn.execute(
                    text("""
                        UPDATE leave_balances SET used = used + :days
                        WHERE user_id=:uid AND leave_type_id=:ltid AND year=:yr
                    """),
                    {"days": lr["days"], "uid": lr["user_id"], "ltid": lr["leave_type_id"], "yr": yr}
                )

    try:
        send_push(lr["user_id"], "Leave Approved ✅", f"Your {lr['leave_type']} request has been approved.")
    except Exception as e:
        print(f"[push warning] {e}")

    return {
        "success": True
    }

@app.post("/leave-requests/{leave_id}/reject")
def reject_leave(leave_id: int, _staff: dict = Depends(require_roles("admin", "manager"))):

    with engine.begin() as conn:

        lr = conn.execute(text("SELECT user_id, leave_type FROM leave_requests WHERE id=:id"), {"id": leave_id}).mappings().first()

        conn.execute(
            text("""
                UPDATE leave_requests
                SET
                    status = 'Rejected',
                    approved_at = NOW()
                WHERE id = :leave_id
            """),
            {
                "leave_id": leave_id
            }
        )

    if lr:
        try:
            send_push(lr["user_id"], "Leave Rejected", f"Your {lr['leave_type']} request was not approved.")
        except Exception as e:
            print(f"[push warning] {e}")

    return {
        "success": True
    }

@app.post("/leave-requests/{leave_id}/escalate")
def escalate_leave(leave_id: int, _staff: dict = Depends(require_roles("admin", "manager"))):
    """
    Manager-only action: forward a (still pending) employee leave request
    straight to admin, just in case that's needed in the moment.
    """
    with engine.begin() as conn:

        admin_row = conn.execute(
            text("SELECT id FROM users WHERE role = 'admin' ORDER BY id LIMIT 1")
        ).first()

        if not admin_row:
            return {
                "success": False,
                "error": "No admin user found to escalate to."
            }

        result = conn.execute(
            text("""
                UPDATE leave_requests
                SET
                    approver_id = :admin_id,
                    escalated = TRUE
                WHERE id = :leave_id
                  AND status = 'Pending'
                RETURNING id
            """),
            {
                "admin_id": admin_row[0],
                "leave_id": leave_id
            }
        )

        updated = result.first()

    if not updated:
        return {
            "success": False,
            "error": "Request not found or no longer pending."
        }

    return {
        "success": True
    }


# ══════════════════════════════════════════════════════════════════
# BUSINESS TRAVEL REQUESTS
# ══════════════════════════════════════════════════════════════════

def _admin_ids(conn):
    rows = conn.execute(text("SELECT id FROM users WHERE LOWER(role) = 'admin'")).fetchall()
    return [r[0] for r in rows]


@app.post("/travel-requests")
def create_travel_request(data: TravelRequestIn):

    try:
        start_d = date.fromisoformat(data.start_date)
        end_d = date.fromisoformat(data.end_date)
    except ValueError:
        return {"success": False, "message": "Invalid date format."}

    if end_d < start_d:
        return {"success": False, "message": "End date must be on or after the start date."}

    with engine.begin() as conn:
        travel_id = conn.execute(
            text("""
                INSERT INTO travel_requests
                (user_id, employee_name, origin, destination, travel_mode,
                 start_date, end_date, purpose, estimated_cost, last_reminder_at)
                VALUES
                (:user_id, :employee_name, :origin, :destination, :travel_mode,
                 :start_date, :end_date, :purpose, :estimated_cost, NOW())
                RETURNING id
            """),
            {
                "user_id": data.user_id,
                "employee_name": data.employee_name,
                "origin": data.origin,
                "destination": data.destination,
                "travel_mode": data.travel_mode,
                "start_date": data.start_date,
                "end_date": data.end_date,
                "purpose": data.purpose,
                "estimated_cost": data.estimated_cost,
            }
        )
        travel_id = travel_id.scalar()
        admins = _admin_ids(conn)

    for admin_id in admins:
        try:
            send_push(
                admin_id,
                "New Travel Request ✈️",
                f"{data.employee_name} requested travel to {data.destination} ({data.start_date} – {data.end_date})."
            )
        except Exception as e:
            print(f"[push warning] {e}")

    return {"success": True, "travel_id": travel_id}


@app.get("/travel-requests")
def get_travel_requests(user_id: Optional[int] = None, status: Optional[str] = None):
    """
    Admin/manager view (all requests) when no user_id is given, or a
    single employee's own requests when user_id is passed. Optional
    status filter, e.g. status=Pending.
    """
    query = "SELECT * FROM travel_requests WHERE 1=1"
    params = {}
    if user_id is not None:
        query += " AND user_id = :user_id"
        params["user_id"] = user_id
    if status is not None:
        query += " AND status = :status"
        params["status"] = status
    query += " ORDER BY created_at DESC"

    with engine.connect() as conn:
        rows = conn.execute(text(query), params).mappings().all()

    return rows


@app.post("/travel-requests/{travel_id}/approve")
def approve_travel_request(travel_id: int, _staff: dict = Depends(require_roles("admin", "ceo"))):

    with engine.begin() as conn:
        tr = conn.execute(text("SELECT * FROM travel_requests WHERE id=:id"), {"id": travel_id}).mappings().first()
        if not tr:
            return {"success": False, "message": "Travel request not found."}

        conn.execute(
            text("""
                UPDATE travel_requests
                SET status = 'Approved', approved_at = NOW()
                WHERE id = :travel_id
            """),
            {"travel_id": travel_id}
        )

    try:
        send_push(tr["user_id"], "Travel Request Approved ✅", f"Your trip to {tr['destination']} has been approved.")
    except Exception as e:
        print(f"[push warning] {e}")

    return {"success": True}


@app.post("/travel-requests/{travel_id}/reject")
def reject_travel_request(travel_id: int, _staff: dict = Depends(require_roles("admin", "ceo"))):

    with engine.begin() as conn:
        tr = conn.execute(text("SELECT * FROM travel_requests WHERE id=:id"), {"id": travel_id}).mappings().first()
        if not tr:
            return {"success": False, "message": "Travel request not found."}

        conn.execute(
            text("""
                UPDATE travel_requests
                SET status = 'Rejected', approved_at = NOW()
                WHERE id = :travel_id
            """),
            {"travel_id": travel_id}
        )

    try:
        send_push(tr["user_id"], "Travel Request Rejected", f"Your trip to {tr['destination']} was not approved.")
    except Exception as e:
        print(f"[push warning] {e}")

    return {"success": True}


@app.post("/travel-requests/{travel_id}/schedule-meeting")
def schedule_travel_meeting(travel_id: int, data: TravelMeetingIn, _staff: dict = Depends(require_roles("admin"))):
    """
    Admin books a discussion meeting about a pending (or already
    approved) travel request. Creates a real entry in the existing
    meetings/calendar system with both the admin and the requester
    as attendees, and links it back onto the travel request.
    """
    with engine.begin() as conn:
        tr = conn.execute(text("SELECT * FROM travel_requests WHERE id=:id"), {"id": travel_id}).mappings().first()
        if not tr:
            return {"success": False, "message": "Travel request not found."}

        conflict = conn.execute(
            text("""
                SELECT id FROM meetings
                WHERE meeting_date = :meeting_date
                AND organizer_id = :organizer_id
                AND (start_slot < :new_end AND end_slot > :new_start)
            """),
            {"meeting_date": data.meeting_date, "organizer_id": _staff["uid"],
             "new_start": data.start_slot, "new_end": data.end_slot}
        ).fetchone()

        if conflict:
            return {"success": False, "message": "You already have a meeting at that time."}

        organizer_id = _staff["uid"]

        meeting_id = conn.execute(
            text("""
                INSERT INTO meetings
                    (title, description, organizer_id, meeting_date, start_slot, end_slot, location)
                VALUES
                    (:title, :description, :organizer_id, :meeting_date, :start_slot, :end_slot, :location)
                RETURNING id
            """),
            {
                "title": f"Travel Request Discussion — {tr['employee_name']}",
                "description": f"Trip: {tr['origin']} → {tr['destination']} ({tr['start_date']} – {tr['end_date']}). Purpose: {tr['purpose']}",
                "organizer_id": organizer_id,
                "meeting_date": data.meeting_date,
                "start_slot": data.start_slot,
                "end_slot": data.end_slot,
                "location": data.location,
            }
        ).scalar()

        for uid in {organizer_id, tr["user_id"]}:
            conn.execute(
                text("INSERT INTO meeting_attendees(meeting_id, user_id) VALUES(:mid, :uid)"),
                {"mid": meeting_id, "uid": uid}
            )

        conn.execute(
            text("UPDATE travel_requests SET meeting_id = :mid WHERE id = :tid"),
            {"mid": meeting_id, "tid": travel_id}
        )

    try:
        send_push(
            tr["user_id"],
            "Meeting Scheduled 📅",
            f"A meeting about your trip to {tr['destination']} was scheduled for {data.meeting_date}."
        )
    except Exception as e:
        print(f"[push warning] {e}")

    return {"success": True, "meeting_id": meeting_id}


@app.post("/cron/travel-request-reminders")
def travel_request_reminders():
    """
    Pushes a reminder to every admin for each travel request that is
    still Pending and hasn't had a reminder sent in the last 4 hours.
    Not triggered automatically — schedule an external cron (Railway
    Cron, cron-job.org, GitHub Actions, etc) to POST here periodically
    (e.g. every 30 minutes); this endpoint enforces the actual 4-hour
    gap itself via last_reminder_at, so calling it more often than
    that is safe and just makes the reminder timing more precise.
    """
    with engine.begin() as conn:
        due = conn.execute(
            text("""
                SELECT * FROM travel_requests
                WHERE status = 'Pending'
                AND (last_reminder_at IS NULL OR last_reminder_at <= NOW() - INTERVAL '4 hours')
            """)
        ).mappings().all()

        admins = _admin_ids(conn)

        sent = 0
        for tr in due:
            for admin_id in admins:
                try:
                    send_push(
                        admin_id,
                        "Travel Request Awaiting Approval ⏳",
                        f"{tr['employee_name']}'s trip to {tr['destination']} is still pending. "
                        f"Approve it or set up a meeting."
                    )
                    sent += 1
                except Exception as e:
                    print(f"[push warning] {e}")
            conn.execute(
                text("UPDATE travel_requests SET last_reminder_at = NOW() WHERE id = :id"),
                {"id": tr["id"]}
            )

    return {"success": True, "reminders_sent": sent, "requests_due": len(due)}


@app.get("/leave-stats")
def leave_stats():

    with engine.connect() as conn:

        total = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM leave_requests
            """)
        ).scalar()

        pending = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM leave_requests
                WHERE status='Pending'
            """)
        ).scalar()

        approved = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM leave_requests
                WHERE status='Approved'
            """)
        ).scalar()

        rejected = conn.execute(
            text("""
                SELECT COUNT(*)
                FROM leave_requests
                WHERE status='Rejected'
            """)
        ).scalar()

    return {
        "total": total,
        "pending": pending,
        "approved": approved,
        "rejected": rejected
    }
@app.get("/leave-requests/pending")
def get_pending_leaves():

    with engine.connect() as conn:

        result = conn.execute(text("""
            SELECT *
            FROM leave_requests
            WHERE status = 'Pending'
            ORDER BY created_at DESC
        """))

        return result.mappings().all()
from sqlalchemy import text

@app.post("/employee-voice")
def submit_employee_voice(data: EmployeeVoiceRequest):

    with engine.begin() as conn:

        result = conn.execute(
            text("""
                INSERT INTO employee_voice
                (
                    user_id,
                    employee_name,
                    is_anonymous,
                    category,
                    priority,
                    subject,
                    description,
                    attachment
                )
                VALUES
                (
                    :user_id,
                    :employee_name,
                    :is_anonymous,
                    :category,
                    :priority,
                    :subject,
                    :description,
                    :attachment
                )
                RETURNING id
            """),
            {
                "user_id": data.user_id,
                "employee_name": data.employee_name,
                "is_anonymous": data.is_anonymous,
                "category": data.category,
                "priority": data.priority,
                "subject": data.subject,
                "description": data.description,
                "attachment": data.attachment
            }
        )

        new_id = result.scalar()

    return {
        "success": True,
        "id": new_id
    }

@app.get("/employee-voice/my/{user_id}")
def get_my_voice(user_id: int):

    with engine.connect() as conn:

        result = conn.execute(
            text("""
                SELECT *
                FROM employee_voice
                WHERE
                    user_id = :user_id
                AND
                    is_anonymous = FALSE
                ORDER BY created_at DESC
            """),
            {
                "user_id": user_id
            }
        )

        return result.mappings().all()

@app.get("/employee-voice")
def get_employee_voice():

    with engine.connect() as conn:

        result = conn.execute(text("""
            SELECT *
            FROM employee_voice
            ORDER BY created_at DESC
        """))

        rows = result.mappings().all()

    response = []

    for row in rows:

        item = dict(row)

        if item["is_anonymous"]:
            item["employee_name"] = "Anonymous"

        response.append(item)

    return response

class VoiceStatusUpdate(BaseModel):
    status: str

@app.post("/employee-voice/{voice_id}/status")
def update_voice_status(
    voice_id: int,
    data: VoiceStatusUpdate
):

    with engine.begin() as conn:

        conn.execute(
            text("""
                UPDATE employee_voice
                SET
                    status = :status,
                    responded_at = NOW()
                WHERE id = :id
            """),
            {
                "status": data.status,
                "id": voice_id
            }
        )

    return {
        "success": True
    }
def _next_serial(conn) -> str:
    """Atomically increments the counter → 5-digit serial e.g. LH-00001."""
    row = conn.execute(
        text("SELECT last_no FROM letterhead_serial_counter WHERE id=1 FOR UPDATE")
    ).fetchone()
    new_no = (row.last_no if row else 0) + 1
    conn.execute(
        text("UPDATE letterhead_serial_counter SET last_no = :n WHERE id=1"),
        {"n": new_no}
    )
    return f"LH-{new_no:05d}"


# ── GET ALL LETTERHEADS ──────────────────────────────────────
@app.get("/letterheads")
def get_letterheads():
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT * FROM letterheads ORDER BY created_at DESC")
        ).mappings().all()
    return [dict(r) for r in rows]


# ── NEXT SERIAL PREVIEW (no increment) ──────────────────────
@app.get("/letterheads/next-serial")
def get_next_serial():
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT last_no FROM letterhead_serial_counter WHERE id=1")
        ).fetchone()
    next_no = (row.last_no if row else 0) + 1
    return {"next_serial": f"LH-{next_no:05d}"}


# ── GENERATE & ISSUE (auto serial) ──────────────────────────
@app.post("/letterheads/generate")
def generate_letterhead(
    department:        str,
    purpose:           str,
    recipient:         str,
    issued_by:         str,
    authorised_by:     str          = "",
    body_content:      str          = "",
    remarks:           str          = "",
    issued_by_user_id: Optional[int] = None,
):
    from datetime import date
    try:
        with engine.begin() as conn:
            serial_no = _next_serial(conn)
            conn.execute(
                text("""
                    INSERT INTO letterheads
                        (serial_no, date_issued, department, purpose,
                         recipient, issued_by, authorised_by, body_content,
                         remarks, issued_by_user_id)
                    VALUES
                        (:serial_no, CURRENT_DATE, :department, :purpose,
                         :recipient, :issued_by, :authorised_by, :body_content,
                         :remarks, :issued_by_user_id)
                """),
                {
                    "serial_no": serial_no, "department": department,
                    "purpose": purpose, "recipient": recipient,
                    "issued_by": issued_by, "authorised_by": authorised_by,
                    "body_content": body_content, "remarks": remarks,
                    "issued_by_user_id": issued_by_user_id,
                }
            )
        return {"success": True, "serial_no": serial_no, "date_issued": str(date.today())}
    except Exception as e:
        return {"success": False, "message": str(e)}


# ── LOG EXISTING PRINTED LETTERHEAD (auto serial) ───────────
@app.post("/letterheads/log")
def log_letterhead(
    date_issued:       str,
    department:        str,
    purpose:           str,
    recipient:         str,
    issued_by:         str,
    authorised_by:     str          = "",
    remarks:           str          = "",
    issued_by_user_id: Optional[int] = None,
):
    """
    Records a pre-printed letterhead used outside the system.
    Serial is auto-assigned from the shared counter so the
    sequence stays continuous across generated and logged entries.
    """
    try:
        with engine.begin() as conn:
            serial_no = _next_serial(conn)
            conn.execute(
                text("""
                    INSERT INTO letterheads
                        (serial_no, date_issued, department, purpose,
                         recipient, issued_by, authorised_by, remarks,
                         issued_by_user_id)
                    VALUES
                        (:serial_no, :date_issued, :department, :purpose,
                         :recipient, :issued_by, :authorised_by, :remarks,
                         :issued_by_user_id)
                """),
                {
                    "serial_no": serial_no, "date_issued": date_issued,
                    "department": department, "purpose": purpose,
                    "recipient": recipient, "issued_by": issued_by,
                    "authorised_by": authorised_by, "remarks": remarks,
                    "issued_by_user_id": issued_by_user_id,
                }
            )
        return {"success": True, "serial_no": serial_no}
    except Exception as e:
        return {"success": False, "message": str(e)}


# ── RESET: wipe all records + reset counter to 0 ────────────
@app.post("/letterheads/reset")
def reset_letterheads(_admin: dict = Depends(require_roles("admin"))):
    """Admin-only. Clears all letterhead records and resets counter to 0."""
    try:
        with engine.begin() as conn:
            conn.execute(text("DELETE FROM letterheads"))
            conn.execute(text("UPDATE letterhead_serial_counter SET last_no = 0 WHERE id = 1"))
        return {"success": True, "message": "All records cleared. Next serial: LH-00001"}
    except Exception as e:
        return {"success": False, "message": str(e)}


# ── VOID A LETTERHEAD ────────────────────────────────────────
@app.post("/letterheads/{letterhead_id}/void")
def void_letterhead(letterhead_id: int, data: VoidRequest, _staff: dict = Depends(require_roles("admin", "manager"))):
    with engine.begin() as conn:
        result = conn.execute(
            text("""
                UPDATE letterheads
                SET status = 'Voided', void_reason = :reason
                WHERE id = :id RETURNING id
            """),
            {"reason": data.void_reason, "id": letterhead_id}
        )
        updated = result.fetchone()
    if not updated:
        return {"success": False, "message": "Letterhead not found."}
    return {"success": True}


# ── GET SINGLE LETTERHEAD ────────────────────────────────────
@app.get("/letterheads/{letterhead_id}")
def get_letterhead(letterhead_id: int):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM letterheads WHERE id = :id"),
            {"id": letterhead_id}
        ).mappings().fetchone()
    if not row:
        return {"success": False, "message": "Not found."}
    return dict(row)


# ── STATS SUMMARY ────────────────────────────────────────────
@app.get("/letterheads/stats/summary")
def letterhead_stats():
    with engine.connect() as conn:
        total      = conn.execute(text("SELECT COUNT(*) FROM letterheads")).scalar()
        voided     = conn.execute(text("SELECT COUNT(*) FROM letterheads WHERE status='Voided'")).scalar()
        this_month = conn.execute(text("""
            SELECT COUNT(*) FROM letterheads
            WHERE DATE_TRUNC('month', date_issued) = DATE_TRUNC('month', CURRENT_DATE)
        """)).scalar()
        by_dept = conn.execute(text("""
            SELECT department, COUNT(*) AS count FROM letterheads
            WHERE status != 'Voided' GROUP BY department
        """)).mappings().all()
        counter_row = conn.execute(
            text("SELECT last_no FROM letterhead_serial_counter WHERE id=1")
        ).fetchone()
    next_no = (counter_row.last_no if counter_row else 0) + 1
    return {
        "total": total, "voided": voided, "this_month": this_month,
        "by_department": [dict(r) for r in by_dept],
        "next_serial": f"LH-{next_no:05d}",
    }

# ================================================================
# PRINT CENTER MODULE
# Organization-wide log of every document printed through the
# system. Any page can call POST /print-logs right before it opens
# the print dialog; the returned id/printed_at is what gets encoded
# into the QR code and printed in the document footer.
# ================================================================

class PrintLogRequest(BaseModel):
    document_title: str = ""
    printed_for:    str
    given_to:       str


# print_logs.printed_at is stored as clock-time IST (not UTC). We deliberately
# compute it in Python rather than relying on the DB's NOW() (which is UTC on
# Railway's Postgres), then always hand it back to clients with an explicit
# "+05:30" suffix so nothing downstream can misinterpret it as UTC.
def _ist_now_naive():
    return datetime.now(ZoneInfo("Asia/Kolkata")).replace(microsecond=0, tzinfo=None)


def _fmt_ist(dt):
    if dt is None:
        return None
    return dt.isoformat() + "+05:30"


# ── CREATE: log a print action (server time + identity are authoritative) ──
@app.post("/print-logs")
def create_print_log(data: PrintLogRequest, user: dict = Depends(get_current_user)):
    try:
        user_id = user["uid"]
        with engine.connect() as conn:
            name_row = conn.execute(
                text("SELECT full_name FROM users WHERE id = :id"),
                {"id": user_id}
            ).fetchone()
        user_name = name_row.full_name if name_row else "User"
        printed_at_ist = _ist_now_naive()

        with engine.begin() as conn:
            row = conn.execute(
                text("""
                    INSERT INTO print_logs
                        (user_id, user_name, document_title, printed_for, given_to, printed_at)
                    VALUES
                        (:user_id, :user_name, :document_title, :printed_for, :given_to, :printed_at)
                    RETURNING id, printed_at
                """),
                {
                    "user_id":        user_id,
                    "user_name":      user_name,
                    "document_title": data.document_title,
                    "printed_for":    data.printed_for,
                    "given_to":       data.given_to,
                    "printed_at":     printed_at_ist,
                }
            ).fetchone()
        return {
            "success": True,
            "id": row.id,
            "printed_at": _fmt_ist(row.printed_at),
            "user_name": user_name,
        }
    except Exception as e:
        return {"success": False, "message": str(e)}


# ── LIST (MINE): every logged-in user's own print history ──────
@app.get("/print-logs/mine")
def get_my_print_logs(user: dict = Depends(get_current_user)):
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT * FROM print_logs
                WHERE user_id = :uid
                ORDER BY printed_at DESC
            """),
            {"uid": user["uid"]}
        ).mappings().all()
    out = []
    for r in rows:
        d = dict(r)
        d["printed_at"] = _fmt_ist(d.get("printed_at"))
        out.append(d)
    return out


# ── LIST: admin-only organization-wide print log ────────────────
@app.get("/print-logs")
def get_print_logs(_admin: dict = Depends(require_roles("admin"))):
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT * FROM print_logs ORDER BY printed_at DESC")
        ).mappings().all()
    out = []
    for r in rows:
        d = dict(r)
        d["printed_at"] = _fmt_ist(d.get("printed_at"))
        out.append(d)
    return out


# ── STATS: admin-only summary cards ─────────────────────────────
@app.get("/print-logs/stats/summary")
def print_log_stats(_admin: dict = Depends(require_roles("admin"))):
    # printed_at is stored as IST clock time, so "today"/"this month" must be
    # computed against IST boundaries, not Postgres's own (UTC) CURRENT_DATE —
    # otherwise counts drift for the first ~5.5 hours of every IST day.
    ist_today       = _ist_now_naive().date()
    ist_month_start = ist_today.replace(day=1)
    with engine.connect() as conn:
        total = conn.execute(text("SELECT COUNT(*) FROM print_logs")).scalar()
        today = conn.execute(text("""
            SELECT COUNT(*) FROM print_logs WHERE DATE(printed_at) = :d
        """), {"d": ist_today}).scalar()
        this_month = conn.execute(text("""
            SELECT COUNT(*) FROM print_logs WHERE printed_at >= :m
        """), {"m": ist_month_start}).scalar()
        by_user = conn.execute(text("""
            SELECT user_name, COUNT(*) AS count FROM print_logs
            GROUP BY user_name ORDER BY count DESC LIMIT 10
        """)).mappings().all()
    return {
        "total": total,
        "today": today,
        "this_month": this_month,
        "by_user": [dict(r) for r in by_user],
    }


# ================================================================
# ATTENDANCE & PAYROLL MODULE
# Paste this entire block at the bottom of your existing main.py
# (before the last closing lines, after the letterhead endpoints)
# ================================================================

from datetime import date, timedelta
from decimal import Decimal


# ── PYDANTIC MODELS ─────────────────────────────────────────────

class AttendanceRecord(BaseModel):
    emp_id: str
    emp_name: str
    department: str = ""
    att_date: str                          # YYYY-MM-DD
    check_in: Optional[str] = None        # HH:MM
    check_out: Optional[str] = None       # HH:MM
    working_hours: Optional[str] = "—"
    status: str = "Present"
    late_minutes: int = 0
    early_leaving: int = 0
    overtime: float = 0.0
    remarks: str = ""
    source: str = "manual"
    created_by: Optional[int] = None


class AttendancePatch(BaseModel):
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    working_hours: Optional[str] = None
    status: Optional[str] = None
    late_minutes: Optional[int] = None
    early_leaving: Optional[int] = None
    overtime: Optional[float] = None
    remarks: Optional[str] = None


class CorrectionRequest(BaseModel):
    emp_id: str
    emp_name: str
    department: str = ""
    att_date: str
    issue_type: str
    req_check_in: Optional[str] = None
    req_check_out: Optional[str] = None
    reason: str
    attachment_url: Optional[str] = None


class CorrectionReview(BaseModel):
    final_check_in: Optional[str] = None
    final_check_out: Optional[str] = None
    admin_notes: str = ""
    reviewed_by: str


class SalaryStructure(BaseModel):
    emp_id: str
    emp_name: str
    basic_salary: float = 0
    hra: float = 0
    special_allowance: float = 0
    travel_allowance: float = 0
    medical_allowance: float = 0
    bonus: float = 0
    incentives: float = 0
    pf_pct: float = 12
    esic_pct: float = 0.75
    professional_tax: float = 200
    income_tax: float = 0
    other_deductions: float = 0
    effective_from: Optional[str] = None
    created_by: Optional[int] = None


class PayrollUnlockRequest(BaseModel):
    unlocked_by: str


class HolidayCreate(BaseModel):
    holiday_date: str
    name: str
    holiday_type: str = "Public"


class AttendanceSettingsUpdate(BaseModel):
    correction_window_hours: Optional[int] = None
    standard_work_hours: Optional[float] = None
    late_grace_minutes: Optional[int] = None
    overtime_after_hours: Optional[float] = None
    office_start_time: Optional[str] = None


# ── HELPERS ─────────────────────────────────────────────────────

def _calc_hours(check_in: Optional[str], check_out: Optional[str]) -> str:
    """Return friendly string e.g. '8h 30m' or '—'."""
    if not check_in or not check_out:
        return "—"
    try:
        fmt = "%H:%M"
        from datetime import datetime as _dt
        ci = _dt.strptime(check_in, fmt)
        co = _dt.strptime(check_out, fmt)
        mins = int((co - ci).total_seconds() / 60)
        if mins <= 0:
            return "—"
        h, m = divmod(mins, 60)
        return f"{h}h {m}m" if m else f"{h}h"
    except Exception:
        return "—"


def _working_days_in_month(year: int, month: int) -> int:
    """Count Mon-Fri days in the given month (no holiday awareness)."""
    import calendar
    total = 0
    for d in range(1, calendar.monthrange(year, month)[1] + 1):
        if date(year, month, d).weekday() < 5:
            total += 1
    return total


def _get_settings(conn) -> dict:
    row = conn.execute(text("SELECT * FROM attendance_settings WHERE id=1")).mappings().fetchone()
    settings = dict(row) if row else {
        "correction_window_hours": 24,
        "standard_work_hours": 9.0,
        "late_grace_minutes": 10,
        "overtime_after_hours": 9.0,
        "office_start_time": "09:00",
    }
    # Postgres TIME columns come back as datetime.time objects, not strings.
    # Normalize to "HH:MM" so downstream .split(":") calls always work.
    ost = settings.get("office_start_time")
    if ost is not None and not isinstance(ost, str):
        settings["office_start_time"] = ost.strftime("%H:%M")
    return settings


# ================================================================
# ATTENDANCE SETTINGS
# ================================================================

@app.get("/attendance/settings")
def get_attendance_settings():
    """Return configurable attendance settings."""
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM attendance_settings WHERE id=1")
        ).mappings().fetchone()
    return dict(row) if row else {}


@app.patch("/attendance/settings")
def update_attendance_settings(data: AttendanceSettingsUpdate, _admin: dict = Depends(require_roles("admin"))):
    """Update one or more attendance settings fields."""
    updates = {k: v for k, v in data.dict().items() if v is not None}
    if not updates:
        return {"success": False, "message": "Nothing to update."}

    set_clause = ", ".join(f"{k} = :{k}" for k in updates)
    updates["updated_at_val"] = datetime.now()

    with engine.begin() as conn:
        conn.execute(
            text(f"UPDATE attendance_settings SET {set_clause}, updated_at = :updated_at_val WHERE id=1"),
            updates
        )
    return {"success": True}


# ================================================================
# ATTENDANCE RECORDS
# ================================================================

@app.get("/attendance")
def get_attendance(
    emp_id: Optional[str]    = None,
    att_date: Optional[str]  = None,
    month: Optional[str]     = None,   # YYYY-MM
    department: Optional[str]= None,
    status: Optional[str]    = None,
    from_date: Optional[str] = None,
    to_date: Optional[str]   = None,
):
    """
    Fetch attendance records with optional filters.
    - emp_id      → single employee
    - att_date    → exact date  (YYYY-MM-DD)
    - month       → all records for that month  (YYYY-MM)
    - department  → filter by dept
    - status      → Present | Absent | …
    - from_date / to_date → date range
    """
    filters = []
    params: dict = {}

    if emp_id:
        filters.append("emp_id = :emp_id")
        params["emp_id"] = emp_id
    if att_date:
        filters.append("att_date = :att_date")
        params["att_date"] = att_date
    if month:
        filters.append("TO_CHAR(att_date,'YYYY-MM') = :month")
        params["month"] = month
    if department:
        filters.append("department = :department")
        params["department"] = department
    if status:
        filters.append("status = :status")
        params["status"] = status
    if from_date:
        filters.append("att_date >= :from_date")
        params["from_date"] = from_date
    if to_date:
        filters.append("att_date <= :to_date")
        params["to_date"] = to_date

    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    sql = f"SELECT * FROM attendance {where} ORDER BY att_date DESC, emp_id"

    with engine.connect() as conn:
        rows = conn.execute(text(sql), params).mappings().all()

    return [dict(r) for r in rows]


@app.post("/attendance")
def create_attendance(data: AttendanceRecord):
    """
    Create a single attendance record.
    Raises HTTP 409 if a record already exists for emp_id + att_date.
    """
    hours = data.working_hours if data.working_hours and data.working_hours != "—" else _calc_hours(data.check_in, data.check_out)

    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    INSERT INTO attendance
                        (emp_id, emp_name, department, att_date,
                         check_in, check_out, working_hours, status,
                         late_minutes, early_leaving, overtime, remarks,
                         source, created_by, updated_at, created_at)
                    VALUES
                        (:emp_id, :emp_name, :department, :att_date,
                         :check_in, :check_out, :working_hours, :status,
                         :late_minutes, :early_leaving, :overtime, :remarks,
                         :source, :created_by, NOW(), NOW())
                    RETURNING id
                """),
                {
                    "emp_id":       data.emp_id,
                    "emp_name":     data.emp_name,
                    "department":   data.department,
                    "att_date":     data.att_date,
                    "check_in":     data.check_in,
                    "check_out":    data.check_out,
                    "working_hours":hours,
                    "status":       data.status,
                    "late_minutes": data.late_minutes,
                    "early_leaving":data.early_leaving,
                    "overtime":     data.overtime,
                    "remarks":      data.remarks,
                    "source":       data.source,
                    "created_by":   data.created_by,
                }
            )
            new_id = result.fetchone()[0]
        return {"success": True, "id": new_id}
    except Exception as e:
        if "unique" in str(e).lower():
            return {"success": False, "message": "Attendance record already exists for this employee on this date."}
        return {"success": False, "message": str(e)}


@app.put("/attendance/{attendance_id}")
def update_attendance(attendance_id: int, data: AttendancePatch, _staff: dict = Depends(require_roles("admin", "manager"))):
    """Full or partial update of an attendance record."""
    updates = {k: v for k, v in data.dict().items() if v is not None}
    if not updates:
        return {"success": False, "message": "Nothing to update."}

    # Recalculate hours if times changed
    if "check_in" in updates or "check_out" in updates:
        with engine.connect() as conn:
            existing = conn.execute(
                text("SELECT check_in, check_out FROM attendance WHERE id=:id"),
                {"id": attendance_id}
            ).fetchone()
        ci = updates.get("check_in", str(existing.check_in) if existing and existing.check_in else None)
        co = updates.get("check_out", str(existing.check_out) if existing and existing.check_out else None)
        updates["working_hours"] = _calc_hours(ci, co)

    updates["updated_at"] = datetime.now()
    set_clause = ", ".join(f"{k} = :{k}" for k in updates)

    with engine.begin() as conn:
        result = conn.execute(
            text(f"UPDATE attendance SET {set_clause} WHERE id=:att_id RETURNING id"),
            {**updates, "att_id": attendance_id}
        )
        row = result.fetchone()

    if not row:
        return {"success": False, "message": "Record not found."}
    return {"success": True}


@app.delete("/attendance/{attendance_id}")
def delete_attendance(attendance_id: int, _admin: dict = Depends(require_roles("admin"))):
    """Delete an attendance record (admin only — enforce role on frontend)."""
    with engine.begin() as conn:
        result = conn.execute(
            text("DELETE FROM attendance WHERE id=:id RETURNING id"),
            {"id": attendance_id}
        )
        row = result.fetchone()
    if not row:
        return {"success": False, "message": "Record not found."}
    return {"success": True}


# ================================================================
# SELF-SERVICE PUNCH IN / PUNCH OUT
# ================================================================

@app.post("/attendance/checkin")
def checkin(_user: dict = Depends(get_current_user)):
    """
    Employee punches in for today.
    Creates attendance row with check_in = current IST time.
    Calculates late_minutes vs office_start_time from attendance_settings.
    emp_id is taken from the verified auth token — never from client input —
    so a logged-in user can only ever punch themselves in.
    """
    emp_id = str(_user["uid"])
    from datetime import datetime as _dt
    now_ist = _dt.now(ZoneInfo("Asia/Kolkata"))
    today   = now_ist.date().isoformat()
    ci_str  = now_ist.strftime("%H:%M")

    with engine.connect() as conn:
        user = conn.execute(
            text("SELECT full_name FROM users WHERE id = :id OR CAST(id AS TEXT) = :ids"),
            {"id": int(emp_id) if emp_id.isdigit() else -1, "ids": emp_id}
        ).fetchone()
        existing = conn.execute(
            text("SELECT id, check_in FROM attendance WHERE emp_id=:eid AND att_date=:d"),
            {"eid": emp_id, "d": today}
        ).fetchone()
        settings = _get_settings(conn)

    office_start = settings.get("office_start_time", "09:00")
    grace        = int(settings.get("late_grace_minutes", 10))
    os_h, os_m   = map(int, office_start.split(":"))
    ci_h, ci_m   = map(int, ci_str.split(":"))
    late_mins    = max(0, (ci_h * 60 + ci_m) - (os_h * 60 + os_m) - grace)
    emp_name     = user.full_name if user else emp_id

    if existing and existing.check_in:
        # Already punched in today (possibly already checked out too).
        # Re-punching in restarts today's record cleanly: reset check_out/hours/overtime
        # explicitly here (rather than relying on a generic PATCH, which ignores null
        # values and would leave a stale check_out behind).
        with engine.begin() as conn:
            conn.execute(
                text("""
                    UPDATE attendance
                    SET check_in=:ci, check_out=NULL, working_hours=NULL, overtime=0,
                        status='Present', late_minutes=:late, updated_at=NOW()
                    WHERE emp_id=:eid AND att_date=:d
                """),
                {"ci": ci_str, "late": late_mins, "eid": emp_id, "d": today}
            )
        return {
            "success": True,
            "check_in": ci_str,
            "late_minutes": late_mins,
            "restarted": True,
            "message": f"Punched in at {ci_str}" + (f" ({late_mins} min late)" if late_mins > 0 else "")
        }

    if existing:
        with engine.begin() as conn:
            conn.execute(
                text("UPDATE attendance SET check_in=:ci, status='Present', late_minutes=:late, updated_at=NOW() WHERE emp_id=:eid AND att_date=:d"),
                {"ci": ci_str, "late": late_mins, "eid": emp_id, "d": today}
            )
    else:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO attendance (emp_id, emp_name, att_date, check_in, status, late_minutes, source, created_at, updated_at)
                    VALUES (:eid, :ename, :d, :ci, 'Present', :late, 'self', NOW(), NOW())
                """),
                {"eid": emp_id, "ename": emp_name, "d": today, "ci": ci_str, "late": late_mins}
            )

    return {
        "success": True,
        "check_in": ci_str,
        "late_minutes": late_mins,
        "message": f"Checked in at {ci_str}" + (f" ({late_mins} min late)" if late_mins > 0 else "")
    }


@app.post("/attendance/checkout")
def checkout(_user: dict = Depends(get_current_user)):
    """
    Employee punches out for today.
    Updates check_out, recalculates working_hours and overtime.
    emp_id is taken from the verified auth token — never from client input.
    """
    emp_id = str(_user["uid"])
    from datetime import datetime as _dt
    now_ist = _dt.now(ZoneInfo("Asia/Kolkata"))
    today   = now_ist.date().isoformat()
    co_str  = now_ist.strftime("%H:%M")

    with engine.connect() as conn:
        rec = conn.execute(
            text("SELECT id, check_in, check_out FROM attendance WHERE emp_id=:eid AND att_date=:d"),
            {"eid": emp_id, "d": today}
        ).fetchone()
        settings = _get_settings(conn)

    if not rec:
        return {"success": False, "message": "No check-in found for today. Please check in first."}
    if rec.check_out:
        return {"success": False, "message": f"Already checked out at {str(rec.check_out)[:5]}"}

    ci_str    = str(rec.check_in)[:5] if rec.check_in else None
    hours_str = _calc_hours(ci_str, co_str)
    std_hours = float(settings.get("standard_work_hours", 9.0))
    ot_hours  = 0.0
    if ci_str:
        ci_h, ci_m = map(int, ci_str.split(":"))
        co_h, co_m = map(int, co_str.split(":"))
        worked_mins = (co_h * 60 + co_m) - (ci_h * 60 + ci_m)
        ot_hours = round(max(0, worked_mins - std_hours * 60) / 60, 2)

    with engine.begin() as conn:
        conn.execute(
            text("UPDATE attendance SET check_out=:co, working_hours=:h, overtime=:ot, updated_at=NOW() WHERE emp_id=:eid AND att_date=:d"),
            {"co": co_str, "h": hours_str, "ot": ot_hours, "eid": emp_id, "d": today}
        )

    return {
        "success": True,
        "check_out": co_str,
        "working_hours": hours_str,
        "overtime": ot_hours,
        "message": f"Checked out at {co_str}. Total: {hours_str}"
    }


@app.get("/attendance/summary/today")
def attendance_summary_today():
    """Quick stat cards for admin dashboard — today's counts."""
    today_str = _ist_today().isoformat()
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT status, COUNT(*) AS cnt FROM attendance WHERE att_date=:d GROUP BY status"),
            {"d": today_str}
        ).mappings().all()
        total = conn.execute(text("SELECT COUNT(*) FROM attendance")).scalar()
        late  = conn.execute(
            text("SELECT COUNT(*) FROM attendance WHERE att_date=:d AND late_minutes > 0"),
            {"d": today_str}
        ).scalar()
    status_map = {r["status"]: r["cnt"] for r in rows}
    return {
        "present": status_map.get("Present", 0) + status_map.get("Work From Home", 0),
        "absent":  status_map.get("Absent", 0),
        "leave":   status_map.get("Leave", 0),
        "late":    late,
        "total_records": total,
    }


@app.get("/attendance/employee/{emp_id}/monthly-summary")
def employee_monthly_summary(emp_id: str, month: str):
    """
    Returns summary counts for an employee for a given month (YYYY-MM).
    Used by the employee My Attendance dashboard cards.
    """
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT status, COUNT(*) AS cnt,
                       SUM(late_minutes) AS total_late,
                       SUM(overtime)     AS total_ot
                FROM attendance
                WHERE emp_id = :emp_id
                  AND TO_CHAR(att_date,'YYYY-MM') = :month
                GROUP BY status
            """),
            {"emp_id": emp_id, "month": month}
        ).mappings().all()

    status_map = {r["status"]: r for r in rows}
    present = (status_map.get("Present", {}).get("cnt") or 0) + \
              (status_map.get("Work From Home", {}).get("cnt") or 0) + \
              (status_map.get("Manual Entry", {}).get("cnt") or 0)
    return {
        "present":    present,
        "absent":     status_map.get("Absent",   {}).get("cnt") or 0,
        "leave":      status_map.get("Leave",    {}).get("cnt") or 0,
        "half_day":   status_map.get("Half Day", {}).get("cnt") or 0,
        "late_days":  sum(1 for r in rows if (r.get("total_late") or 0) > 0),
        "overtime_hours": float(sum((r.get("total_ot") or 0) for r in rows)),
    }


# ================================================================
# ATTENDANCE CORRECTION REQUESTS
# ================================================================

@app.get("/attendance/corrections")
def get_corrections(
    emp_id: Optional[str]     = None,
    status: Optional[str]     = None,
    department: Optional[str] = None,
    att_date: Optional[str]   = None,
):
    """List correction requests. HR/Admin see all; filter by emp_id for self-service."""
    filters, params = [], {}
    if emp_id:
        filters.append("emp_id = :emp_id"); params["emp_id"] = emp_id
    if status:
        filters.append("status = :status"); params["status"] = status
    if department:
        filters.append("department = :department"); params["department"] = department
    if att_date:
        filters.append("att_date = :att_date"); params["att_date"] = att_date

    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT * FROM attendance_corrections {where} ORDER BY submitted_at DESC"),
            params
        ).mappings().all()
    return [dict(r) for r in rows]


@app.get("/attendance/corrections/{correction_id}")
def get_correction_detail(correction_id: int):
    """Full detail including timeline for a single correction request."""
    with engine.connect() as conn:
        corr = conn.execute(
            text("SELECT * FROM attendance_corrections WHERE id=:id"),
            {"id": correction_id}
        ).mappings().fetchone()
        if not corr:
            return {"success": False, "message": "Not found."}

        timeline = conn.execute(
            text("SELECT * FROM correction_timeline WHERE correction_id=:id ORDER BY created_at"),
            {"id": correction_id}
        ).mappings().all()

        # Original attendance record
        att = conn.execute(
            text("SELECT * FROM attendance WHERE emp_id=:eid AND att_date=:d"),
            {"eid": corr["emp_id"], "d": corr["att_date"]}
        ).mappings().fetchone()

    return {
        **dict(corr),
        "timeline": [dict(t) for t in timeline],
        "original_attendance": dict(att) if att else None,
    }


@app.post("/attendance/corrections")
def submit_correction(data: CorrectionRequest):
    """
    Employee submits a correction request.
    Enforces the configurable time window (default 24 hours).
    """
    with engine.connect() as conn:
        settings = _get_settings(conn)
        # Check for duplicate pending request
        existing = conn.execute(
            text("""
                SELECT id FROM attendance_corrections
                WHERE emp_id=:eid AND att_date=:d AND status='Pending'
            """),
            {"eid": data.emp_id, "d": data.att_date}
        ).fetchone()

    if existing:
        return {"success": False, "message": "A pending correction request already exists for this date."}

    # Enforce time window
    att_date_obj = date.fromisoformat(data.att_date)
    diff_hours = (_ist_today() - att_date_obj).days * 24
    window = settings.get("correction_window_hours", 24)
    if diff_hours > window:
        return {
            "success": False,
            "message": f"Attendance correction requests can only be submitted within {window} hours. Please contact HR."
        }

    with engine.begin() as conn:
        # Get linked attendance id
        att_row = conn.execute(
            text("SELECT id FROM attendance WHERE emp_id=:eid AND att_date=:d"),
            {"eid": data.emp_id, "d": data.att_date}
        ).fetchone()

        result = conn.execute(
            text("""
                INSERT INTO attendance_corrections
                    (emp_id, emp_name, department, att_date, issue_type,
                     req_check_in, req_check_out, reason, attachment_url,
                     status, submitted_at, attendance_id)
                VALUES
                    (:emp_id, :emp_name, :department, :att_date, :issue_type,
                     :req_check_in, :req_check_out, :reason, :attachment_url,
                     'Pending', NOW(), :attendance_id)
                RETURNING id
            """),
            {
                "emp_id":         data.emp_id,
                "emp_name":       data.emp_name,
                "department":     data.department,
                "att_date":       data.att_date,
                "issue_type":     data.issue_type,
                "req_check_in":   data.req_check_in,
                "req_check_out":  data.req_check_out,
                "reason":         data.reason,
                "attachment_url": data.attachment_url,
                "attendance_id":  att_row[0] if att_row else None,
            }
        )
        new_id = result.fetchone()[0]

        # Timeline event
        conn.execute(
            text("""
                INSERT INTO correction_timeline (correction_id, event, done_by, notes)
                VALUES (:cid, 'Submitted', :by, '')
            """),
            {"cid": new_id, "by": data.emp_name}
        )

    return {"success": True, "id": new_id}


@app.post("/attendance/corrections/{correction_id}/approve")
def approve_correction(correction_id: int, data: CorrectionReview, _staff: dict = Depends(require_roles("admin", "manager"))):
    """
    HR/Admin approves a correction request.
    - Updates the attendance record with final times.
    - Creates an audit log entry.
    - Warns if payroll is already locked for that month.
    """
    with engine.connect() as conn:
        corr = conn.execute(
            text("SELECT * FROM attendance_corrections WHERE id=:id"),
            {"id": correction_id}
        ).mappings().fetchone()

    if not corr:
        return {"success": False, "message": "Request not found."}
    if corr["status"] != "Pending":
        return {"success": False, "message": f"Request is already {corr['status']}."}

    payroll_month = str(corr["att_date"])[:7]   # YYYY-MM (att_date may be a date object or string)
    payroll_locked = False

    with engine.begin() as conn:
        # Check payroll lock
        lock = conn.execute(
            text("SELECT id FROM payroll_locks WHERE payroll_month=:m AND unlocked_at IS NULL"),
            {"m": payroll_month}
        ).fetchone()
        payroll_locked = lock is not None

        # Fetch original attendance
        att = conn.execute(
            text("SELECT * FROM attendance WHERE emp_id=:eid AND att_date=:d"),
            {"eid": corr["emp_id"], "d": corr["att_date"]}
        ).mappings().fetchone()

        # Strip seconds from TIME objects returned by PostgreSQL ("HH:MM:SS" → "HH:MM")
        def _t(val):
            return str(val)[:5] if val else None

        orig_in  = _t(att["check_in"])  if att else None
        orig_out = _t(att["check_out"]) if att else None

        final_ci = data.final_check_in  or (_t(att["check_in"])  if att else None)
        final_co = data.final_check_out or (_t(att["check_out"]) if att else None)
        new_hours = _calc_hours(final_ci, final_co)

        # Update attendance
        if att:
            conn.execute(
                text("""
                    UPDATE attendance
                    SET check_in=:ci, check_out=:co, working_hours=:h,
                        status='Manual Entry', updated_at=NOW()
                    WHERE emp_id=:eid AND att_date=:d
                """),
                {"ci": final_ci, "co": final_co, "h": new_hours,
                 "eid": corr["emp_id"], "d": corr["att_date"]}
            )
        else:
            # No existing record — create one
            conn.execute(
                text("""
                    INSERT INTO attendance
                        (emp_id, emp_name, department, att_date, check_in, check_out,
                         working_hours, status, source, updated_at, created_at)
                    VALUES
                        (:eid, :ename, :dept, :d, :ci, :co,
                         :h, 'Manual Entry', 'correction', NOW(), NOW())
                """),
                {"eid": corr["emp_id"], "ename": corr["emp_name"],
                 "dept": corr["department"], "d": corr["att_date"],
                 "ci": final_ci, "co": final_co, "h": new_hours}
            )

        # Audit log
        conn.execute(
            text("""
                INSERT INTO attendance_audit
                    (attendance_id, correction_id, emp_id, emp_name, att_date,
                     orig_check_in, orig_check_out, new_check_in, new_check_out,
                     change_reason, approved_by, approved_at)
                VALUES
                    (:att_id, :cid, :eid, :ename, :d,
                     :orig_in, :orig_out, :new_in, :new_out,
                     :reason, :by, NOW())
            """),
            {
                "att_id":   corr.get("attendance_id"),
                "cid":      correction_id,
                "eid":      corr["emp_id"],
                "ename":    corr["emp_name"],
                "d":        corr["att_date"],
                "orig_in":  orig_in,
                "orig_out": orig_out,
                "new_in":   final_ci,
                "new_out":  final_co,
                "reason":   corr["reason"],
                "by":       data.reviewed_by,
            }
        )

        # Update correction status
        conn.execute(
            text("""
                UPDATE attendance_corrections
                SET status='Approved', reviewed_by=:by, reviewed_at=NOW(), admin_notes=:notes
                WHERE id=:id
            """),
            {"by": data.reviewed_by, "notes": data.admin_notes, "id": correction_id}
        )

        # Timeline
        conn.execute(
            text("""
                INSERT INTO correction_timeline (correction_id, event, done_by, notes)
                VALUES (:cid, 'Approved', :by, :notes)
            """),
            {"cid": correction_id, "by": data.reviewed_by, "notes": data.admin_notes}
        )

    return {
        "success": True,
        "payroll_locked_warning": payroll_locked,
        "message": "Approved. Payroll recalculation is recommended." if payroll_locked else "Approved."
    }


@app.post("/attendance/corrections/{correction_id}/reject")
def reject_correction(correction_id: int, data: CorrectionReview, _staff: dict = Depends(require_roles("admin", "manager"))):
    """HR/Admin rejects a correction request."""
    with engine.connect() as conn:
        corr = conn.execute(
            text("SELECT status FROM attendance_corrections WHERE id=:id"),
            {"id": correction_id}
        ).fetchone()

    if not corr:
        return {"success": False, "message": "Request not found."}
    if corr[0] != "Pending":
        return {"success": False, "message": f"Request is already {corr[0]}."}

    with engine.begin() as conn:
        conn.execute(
            text("""
                UPDATE attendance_corrections
                SET status='Rejected', reviewed_by=:by, reviewed_at=NOW(), admin_notes=:notes
                WHERE id=:id
            """),
            {"by": data.reviewed_by, "notes": data.admin_notes, "id": correction_id}
        )
        conn.execute(
            text("""
                INSERT INTO correction_timeline (correction_id, event, done_by, notes)
                VALUES (:cid, 'Rejected', :by, :notes)
            """),
            {"cid": correction_id, "by": data.reviewed_by, "notes": data.admin_notes}
        )
    return {"success": True}


@app.post("/attendance/corrections/{correction_id}/cancel")
def cancel_correction(correction_id: int, emp_name: str):
    """Employee cancels their own pending request."""
    with engine.begin() as conn:
        result = conn.execute(
            text("""
                UPDATE attendance_corrections
                SET status='Cancelled'
                WHERE id=:id AND status='Pending'
                RETURNING id
            """),
            {"id": correction_id}
        )
        row = result.fetchone()
        if row:
            conn.execute(
                text("""
                    INSERT INTO correction_timeline (correction_id, event, done_by)
                    VALUES (:cid, 'Cancelled', :by)
                """),
                {"cid": correction_id, "by": emp_name}
            )
    if not row:
        return {"success": False, "message": "Request not found or already reviewed."}
    return {"success": True}


@app.get("/attendance/corrections/summary/counts")
def corrections_summary():
    """Quick counts for the admin cards."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT status, COUNT(*) AS cnt FROM attendance_corrections GROUP BY status")
        ).mappings().all()
    m = {r["status"]: r["cnt"] for r in rows}
    return {
        "pending":  m.get("Pending", 0),
        "approved": m.get("Approved", 0),
        "rejected": m.get("Rejected", 0),
        "cancelled":m.get("Cancelled", 0),
    }


# ================================================================
# AUDIT LOG
# ================================================================

@app.get("/attendance/audit")
def get_audit_log(
    emp_id: Optional[str]    = None,
    from_date: Optional[str] = None,
    to_date: Optional[str]   = None,
    search: Optional[str]    = None,
):
    filters, params = [], {}
    if emp_id:
        filters.append("emp_id = :emp_id"); params["emp_id"] = emp_id
    if from_date:
        filters.append("att_date >= :from_date"); params["from_date"] = from_date
    if to_date:
        filters.append("att_date <= :to_date"); params["to_date"] = to_date
    if search:
        filters.append("(LOWER(emp_name) LIKE :s OR LOWER(emp_id) LIKE :s)")
        params["s"] = f"%{search.lower()}%"

    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT * FROM attendance_audit {where} ORDER BY approved_at DESC"),
            params
        ).mappings().all()
    return [dict(r) for r in rows]


# ================================================================
# SALARY STRUCTURE
# ================================================================

@app.get("/salary-structure")
def get_all_salary_structures():
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT * FROM salary_structure ORDER BY emp_id")
        ).mappings().all()
    return [dict(r) for r in rows]


@app.get("/salary-structure/{emp_id}")
def get_salary_structure(emp_id: str):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM salary_structure WHERE emp_id=:eid"),
            {"eid": emp_id}
        ).mappings().fetchone()
    if not row:
        return {"success": False, "message": "No salary structure found for this employee."}
    return dict(row)


@app.post("/salary-structure")
def upsert_salary_structure(data: SalaryStructure, _admin: dict = Depends(require_roles("admin"))):
    """Create or update the salary structure for an employee."""
    eff = data.effective_from or date.today().isoformat()
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO salary_structure
                    (emp_id, emp_name, basic_salary, hra, special_allowance,
                     travel_allowance, medical_allowance, bonus, incentives,
                     pf_pct, esic_pct, professional_tax, income_tax,
                     other_deductions, effective_from, created_by, updated_at)
                VALUES
                    (:emp_id, :emp_name, :basic_salary, :hra, :special_allowance,
                     :travel_allowance, :medical_allowance, :bonus, :incentives,
                     :pf_pct, :esic_pct, :professional_tax, :income_tax,
                     :other_deductions, :effective_from, :created_by, NOW())
                ON CONFLICT (emp_id) DO UPDATE SET
                    emp_name          = EXCLUDED.emp_name,
                    basic_salary      = EXCLUDED.basic_salary,
                    hra               = EXCLUDED.hra,
                    special_allowance = EXCLUDED.special_allowance,
                    travel_allowance  = EXCLUDED.travel_allowance,
                    medical_allowance = EXCLUDED.medical_allowance,
                    bonus             = EXCLUDED.bonus,
                    incentives        = EXCLUDED.incentives,
                    pf_pct            = EXCLUDED.pf_pct,
                    esic_pct          = EXCLUDED.esic_pct,
                    professional_tax  = EXCLUDED.professional_tax,
                    income_tax        = EXCLUDED.income_tax,
                    other_deductions  = EXCLUDED.other_deductions,
                    effective_from    = EXCLUDED.effective_from,
                    updated_at        = NOW()
            """),
            {
                "emp_id":           data.emp_id,
                "emp_name":         data.emp_name,
                "basic_salary":     data.basic_salary,
                "hra":              data.hra,
                "special_allowance":data.special_allowance,
                "travel_allowance": data.travel_allowance,
                "medical_allowance":data.medical_allowance,
                "bonus":            data.bonus,
                "incentives":       data.incentives,
                "pf_pct":           data.pf_pct,
                "esic_pct":         data.esic_pct,
                "professional_tax": data.professional_tax,
                "income_tax":       data.income_tax,
                "other_deductions": data.other_deductions,
                "effective_from":   eff,
                "created_by":       data.created_by,
            }
        )
    return {"success": True}


# ================================================================
# PAYROLL
# ================================================================

@app.get("/payroll")
def get_payroll(month: Optional[str] = None, emp_id: Optional[str] = None):
    filters, params = [], {}
    if month:
        filters.append("payroll_month=:month"); params["month"] = month
    if emp_id:
        filters.append("emp_id=:emp_id"); params["emp_id"] = emp_id
    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT * FROM payroll {where} ORDER BY emp_id"),
            params
        ).mappings().all()
    return [dict(r) for r in rows]


@app.post("/payroll/generate")
def generate_payroll(month: str, generated_by: str = "HR Admin", _admin: dict = Depends(require_roles("admin"))):
    """
    Generates payroll for ALL employees who have a salary structure,
    for the given month (YYYY-MM).
    Fails if payroll is already locked for that month.
    Overwrites any un-locked draft for the same month.
    """
    # Validate month format
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except Exception:
        return {"success": False, "message": "Invalid month format. Use YYYY-MM."}

    with engine.connect() as conn:
        lock = conn.execute(
            text("SELECT id FROM payroll_locks WHERE payroll_month=:m AND unlocked_at IS NULL"),
            {"m": month}
        ).fetchone()
        if lock:
            return {"success": False, "message": "Payroll is locked for this month. Unlock first."}

        # All salary structures
        structs = conn.execute(
            text("SELECT * FROM salary_structure")
        ).mappings().all()

        settings = _get_settings(conn)

    working_days = _working_days_in_month(year, mon)
    generated_rows = []

    for ss in structs:
        emp_id_val = ss["emp_id"]

        with engine.connect() as conn:
            att_rows = conn.execute(
                text("""
                    SELECT status, late_minutes, overtime
                    FROM attendance
                    WHERE emp_id=:eid
                      AND TO_CHAR(att_date,'YYYY-MM')=:month
                """),
                {"eid": emp_id_val, "month": month}
            ).mappings().all()

        present    = sum(1 for a in att_rows if a["status"] in ("Present","Work From Home","Manual Entry"))
        half_days  = sum(1 for a in att_rows if a["status"] == "Half Day")
        leave_days = sum(1 for a in att_rows if a["status"] == "Leave")
        ot_hours   = float(sum(a["overtime"] or 0 for a in att_rows))

        effective_present = present + half_days * 0.5 + leave_days
        lop = max(0.0, working_days - effective_present)

        # Earnings
        basic  = float(ss["basic_salary"])
        hra    = float(ss["hra"])
        spl    = float(ss["special_allowance"])
        travel = float(ss["travel_allowance"])
        med    = float(ss["medical_allowance"])
        bonus  = float(ss["bonus"])
        inc    = float(ss["incentives"])
        gross  = basic + hra + spl + travel + med + bonus + inc

        # Deductions
        lop_ded   = round((basic / working_days) * lop, 2) if working_days > 0 else 0
        pf_ded    = round(basic * float(ss["pf_pct"]) / 100, 2)
        esic_ded  = round(gross * float(ss["esic_pct"]) / 100, 2)
        pt_ded    = float(ss["professional_tax"])
        it_ded    = float(ss["income_tax"])
        other_ded = float(ss["other_deductions"])
        total_ded = pf_ded + esic_ded + pt_ded + it_ded + other_ded + lop_ded
        net       = gross - total_ded

        generated_rows.append({
            "month": month, "emp_id": emp_id_val,
            "emp_name": ss["emp_name"], "dept": "",
            "working_days": working_days, "present": present,
            "half_days": half_days, "leave": leave_days, "lop": lop,
            "ot_hours": ot_hours, "gross": gross,
            "pf": pf_ded, "esic": esic_ded, "pt": pt_ded, "it": it_ded,
            "other": other_ded, "lop_ded": lop_ded,
            "total_ded": total_ded, "net": net,
        })

    with engine.begin() as conn:
        # Wipe old draft for this month
        conn.execute(text("DELETE FROM payroll WHERE payroll_month=:m"), {"m": month})
        for r in generated_rows:
            conn.execute(
                text("""
                    INSERT INTO payroll
                        (payroll_month, emp_id, emp_name, department,
                         working_days, present_days, half_days, leave_days, lop_days,
                         overtime_hours, gross_salary,
                         pf_deduction, esic_deduction, pt_deduction, it_deduction,
                         other_deductions, lop_deduction, total_deductions, net_salary,
                         status, generated_at)
                    VALUES
                        (:m, :eid, :ename, :dept,
                         :wd, :p, :hd, :l, :lop,
                         :ot, :gross,
                         :pf, :esic, :pt, :it,
                         :other, :lop_ded, :total_ded, :net,
                         'Generated', NOW())
                """),
                {
                    "m": r["month"], "eid": r["emp_id"], "ename": r["emp_name"], "dept": r["dept"],
                    "wd": r["working_days"], "p": r["present"], "hd": r["half_days"],
                    "l": r["leave"], "lop": r["lop"], "ot": r["ot_hours"],
                    "gross": r["gross"], "pf": r["pf"], "esic": r["esic"],
                    "pt": r["pt"], "it": r["it"], "other": r["other"],
                    "lop_ded": r["lop_ded"], "total_ded": r["total_ded"], "net": r["net"],
                }
            )

    total_gross = sum(r["gross"] for r in generated_rows)
    total_net   = sum(r["net"]   for r in generated_rows)
    return {
        "success": True,
        "employees_processed": len(generated_rows),
        "total_gross": round(total_gross, 2),
        "total_net": round(total_net, 2),
    }


@app.post("/payroll/lock")
def lock_payroll(month: str, locked_by: str = "HR Admin", _admin: dict = Depends(require_roles("admin"))):
    """
    Locks payroll for a month and auto-generates payslips for all employees.
    After locking, no edits to payroll or attendance corrections are applied
    to the payroll figures without an explicit unlock.
    """
    with engine.connect() as conn:
        existing_lock = conn.execute(
            text("SELECT id FROM payroll_locks WHERE payroll_month=:m AND unlocked_at IS NULL"),
            {"m": month}
        ).fetchone()
        if existing_lock:
            return {"success": False, "message": "Payroll is already locked for this month."}

        payroll_rows = conn.execute(
            text("""
                SELECT
                    p.id AS payroll_id, p.payroll_month, p.emp_id, p.emp_name, p.department,
                    p.working_days, p.present_days, p.leave_days, p.lop_days, p.overtime_hours,
                    p.gross_salary, p.pf_deduction, p.esic_deduction, p.pt_deduction, p.it_deduction,
                    p.other_deductions, p.lop_deduction, p.total_deductions, p.net_salary,
                    s.basic_salary, s.hra, s.special_allowance, s.travel_allowance,
                    s.medical_allowance, s.bonus, s.incentives
                FROM payroll p
                LEFT JOIN salary_structure s ON p.emp_id = s.emp_id
                WHERE p.payroll_month=:m
            """),
            {"m": month}
        ).mappings().all()

    if not payroll_rows:
        return {"success": False, "message": "No payroll data found. Generate payroll first."}

    mon_num = int(month[5:7])
    yr_num  = int(month[:4])
    mon_str = f"{mon_num:02d}"

    with engine.begin() as conn:
        # Insert lock record
        conn.execute(
            text("""
                INSERT INTO payroll_locks (payroll_month, locked_by, locked_at)
                VALUES (:m, :by, NOW())
                ON CONFLICT (payroll_month) DO UPDATE SET locked_by=EXCLUDED.locked_by, locked_at=NOW(), unlocked_at=NULL
            """),
            {"m": month, "by": locked_by}
        )

        # Update all payroll rows to Locked
        conn.execute(
            text("UPDATE payroll SET status='Locked' WHERE payroll_month=:m"),
            {"m": month}
        )

        # Generate payslips
        conn.execute(text("DELETE FROM payslips WHERE payroll_month=:m"), {"m": month})
        for row in payroll_rows:
            payslip_no = f"PNT-{yr_num}{mon_str}-{row['emp_id']}"
            conn.execute(
                text("""
                    INSERT INTO payslips
                        (payslip_no, payroll_month, emp_id, emp_name, department, designation,
                         working_days, present_days, leave_days, lop_days, overtime_hours,
                         basic_salary, hra, special_allowance, travel_allowance,
                         medical_allowance, bonus, incentives, gross_salary,
                         pf_deduction, esic_deduction, pt_deduction, it_deduction,
                         other_deductions, lop_deduction, total_deductions, net_salary,
                         generated_at, payroll_id)
                    VALUES
                        (:pno, :m, :eid, :ename, :dept, :desig,
                         :wd, :p, :l, :lop, :ot,
                         :basic, :hra, :spl, :travel, :med, :bonus, :inc, :gross,
                         :pf, :esic, :pt, :it, :other, :lop_ded, :total_ded, :net,
                         NOW(), :pr_id)
                """),
                {
                    "pno":   payslip_no,
                    "m":     month,
                    "eid":   row["emp_id"],
                    "ename": row["emp_name"],
                    "dept":  row.get("department",""),
                    "desig": "Team Member",
                    "wd":    row["working_days"],
                    "p":     row["present_days"],
                    "l":     row["leave_days"],
                    "lop":   row["lop_days"],
                    "ot":    row["overtime_hours"],
                    "basic": row.get("basic_salary", 0),
                    "hra":   row.get("hra", 0),
                    "spl":   row.get("special_allowance", 0),
                    "travel":row.get("travel_allowance", 0),
                    "med":   row.get("medical_allowance", 0),
                    "bonus": row.get("bonus", 0),
                    "inc":   row.get("incentives", 0),
                    "gross": row["gross_salary"],
                    "pf":    row["pf_deduction"],
                    "esic":  row["esic_deduction"],
                    "pt":    row["pt_deduction"],
                    "it":    row["it_deduction"],
                    "other": row["other_deductions"],
                    "lop_ded": row["lop_deduction"],
                    "total_ded": row["total_deductions"],
                    "net":   row["net_salary"],
                    "pr_id": row["payroll_id"],
                }
            )

        # Audit the lock
        conn.execute(
            text("""
                INSERT INTO attendance_audit
                    (emp_id, emp_name, att_date, orig_check_in, orig_check_out,
                     new_check_in, new_check_out, change_reason, approved_by, approved_at)
                VALUES
                    ('SYSTEM','System',:d,'UNLOCKED','UNLOCKED','LOCKED','LOCKED',
                     'Payroll locked by HR', :by, NOW())
            """),
            {"d": f"{yr_num}-{mon_str}-01", "by": locked_by}
        )

    for row in payroll_rows:
        try:
            send_push(row["emp_id"], "Payslip Ready 💰", f"Your payslip for {month} has been generated.")
        except Exception as e:
            print(f"[push warning] {e}")

    return {"success": True, "payslips_generated": len(payroll_rows)}


@app.post("/payroll/unlock")
def unlock_payroll(month: str, data: PayrollUnlockRequest, _admin: dict = Depends(require_roles("admin", "ceo"))):
    """Super Admin unlocks a previously locked payroll month."""
    with engine.begin() as conn:
        result = conn.execute(
            text("""
                UPDATE payroll_locks
                SET unlocked_by=:by, unlocked_at=NOW()
                WHERE payroll_month=:m AND unlocked_at IS NULL
                RETURNING id
            """),
            {"m": month, "by": data.unlocked_by}
        )
        row = result.fetchone()
        if not row:
            return {"success": False, "message": "Payroll is not locked or already unlocked."}

        conn.execute(
            text("UPDATE payroll SET status='Generated' WHERE payroll_month=:m"),
            {"m": month}
        )

        # Audit
        conn.execute(
            text("""
                INSERT INTO attendance_audit
                    (emp_id, emp_name, att_date, orig_check_in, orig_check_out,
                     new_check_in, new_check_out, change_reason, approved_by, approved_at)
                VALUES
                    ('SYSTEM','System',:d,'LOCKED','LOCKED','UNLOCKED','UNLOCKED',
                     'Payroll UNLOCKED by Super Admin', :by, NOW())
            """),
            {"d": f"{month}-01", "by": data.unlocked_by}
        )

    return {"success": True}


@app.get("/payroll/lock-status/{month}")
def payroll_lock_status(month: str):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM payroll_locks WHERE payroll_month=:m"),
            {"m": month}
        ).mappings().fetchone()
    if not row:
        return {"locked": False}
    return {**dict(row), "locked": row["unlocked_at"] is None}


@app.get("/payroll/summary/{month}")
def payroll_summary(month: str):
    """Aggregate totals for a payroll month — used by report cards."""
    with engine.connect() as conn:
        row = conn.execute(
            text("""
                SELECT COUNT(*) AS employees,
                       SUM(gross_salary)    AS gross,
                       SUM(total_deductions)AS deductions,
                       SUM(net_salary)      AS net,
                       SUM(CASE WHEN lop_days > 0 THEN 1 ELSE 0 END) AS lop_employees
                FROM payroll
                WHERE payroll_month=:m
            """),
            {"m": month}
        ).mappings().fetchone()
        lock = conn.execute(
            text("SELECT locked_at, locked_by FROM payroll_locks WHERE payroll_month=:m AND unlocked_at IS NULL"),
            {"m": month}
        ).fetchone()
    return {
        "month": month,
        "employees":     row["employees"] or 0,
        "gross":         float(row["gross"] or 0),
        "deductions":    float(row["deductions"] or 0),
        "net":           float(row["net"] or 0),
        "lop_employees": row["lop_employees"] or 0,
        "locked":        lock is not None,
        "locked_at":     str(lock[0]) if lock else None,
        "locked_by":     lock[1] if lock else None,
    }


# ================================================================
# PAYSLIPS
# ================================================================

@app.get("/payslips")
def get_payslips(
    emp_id: Optional[str] = None,
    month: Optional[str]  = None,
    year: Optional[str]   = None,
):
    filters, params = [], {}
    if emp_id:
        filters.append("emp_id=:emp_id"); params["emp_id"] = emp_id
    if month:
        filters.append("payroll_month=:month"); params["month"] = month
    if year:
        filters.append("payroll_month LIKE :year"); params["year"] = f"{year}-%"
    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT * FROM payslips {where} ORDER BY payroll_month DESC"),
            params
        ).mappings().all()
    return [dict(r) for r in rows]


@app.get("/payslips/{payslip_id}")
def get_payslip(payslip_id: int):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM payslips WHERE id=:id"),
            {"id": payslip_id}
        ).mappings().fetchone()
    if not row:
        return {"success": False, "message": "Payslip not found."}
    return dict(row)


@app.get("/payslips/by-no/{payslip_no}")
def get_payslip_by_no(payslip_no: str):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT * FROM payslips WHERE payslip_no=:no"),
            {"no": payslip_no}
        ).mappings().fetchone()
    if not row:
        return {"success": False, "message": "Payslip not found."}
    return dict(row)


# ================================================================
# HOLIDAYS
# ================================================================

@app.get("/holidays")
def get_holidays(year: Optional[str] = None):
    filters, params = [], {}
    if year:
        filters.append("EXTRACT(YEAR FROM holiday_date)=:year"); params["year"] = int(year)
    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT * FROM holidays {where} ORDER BY holiday_date"),
            params
        ).mappings().all()
    return [dict(r) for r in rows]


@app.post("/holidays")
def create_holiday(data: HolidayCreate, _admin: dict = Depends(require_roles("admin"))):
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    INSERT INTO holidays (holiday_date, name, holiday_type)
                    VALUES (:d, :name, :type)
                    RETURNING id
                """),
                {"d": data.holiday_date, "name": data.name, "type": data.holiday_type}
            )
            new_id = result.fetchone()[0]
        return {"success": True, "id": new_id}
    except Exception as e:
        if "unique" in str(e).lower():
            return {"success": False, "message": "A holiday already exists on this date."}
        return {"success": False, "message": str(e)}


@app.delete("/holidays/{holiday_id}")
def delete_holiday(holiday_id: int, _admin: dict = Depends(require_roles("admin"))):
    with engine.begin() as conn:
        result = conn.execute(
            text("DELETE FROM holidays WHERE id=:id RETURNING id"),
            {"id": holiday_id}
        )
        row = result.fetchone()
    if not row:
        return {"success": False, "message": "Holiday not found."}
    return {"success": True}


# ================================================================
# REPORTS
# ================================================================

@app.get("/reports/attendance/late")
def report_late_employees(month: str):
    """Employees with late arrivals in a given month."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT emp_id, emp_name, department,
                       COUNT(*) AS late_days,
                       SUM(late_minutes) AS total_late_minutes,
                       ROUND(AVG(late_minutes),0) AS avg_late_minutes
                FROM attendance
                WHERE TO_CHAR(att_date,'YYYY-MM')=:month
                  AND late_minutes > 0
                GROUP BY emp_id, emp_name, department
                ORDER BY total_late_minutes DESC
            """),
            {"month": month}
        ).mappings().all()
    return [dict(r) for r in rows]


@app.get("/reports/attendance/absentees")
def report_absentees(month: str):
    """Absent / leave / LOP days per employee for a given month."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT emp_id, emp_name, department,
                       SUM(CASE WHEN status='Absent' THEN 1 ELSE 0 END) AS absent_days,
                       SUM(CASE WHEN status='Leave'  THEN 1 ELSE 0 END) AS leave_days
                FROM attendance
                WHERE TO_CHAR(att_date,'YYYY-MM')=:month
                GROUP BY emp_id, emp_name, department
                HAVING SUM(CASE WHEN status IN ('Absent','Leave') THEN 1 ELSE 0 END) > 0
                ORDER BY absent_days DESC
            """),
            {"month": month}
        ).mappings().all()
    return [dict(r) for r in rows]


@app.get("/reports/attendance/overtime")
def report_overtime(month: str):
    """Overtime summary per employee for a given month."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT emp_id, emp_name, department,
                       COUNT(*) AS ot_days,
                       SUM(overtime) AS ot_hours
                FROM attendance
                WHERE TO_CHAR(att_date,'YYYY-MM')=:month
                  AND overtime > 0
                GROUP BY emp_id, emp_name, department
                ORDER BY ot_hours DESC
            """),
            {"month": month}
        ).mappings().all()
    return [dict(r) for r in rows]


@app.get("/reports/payroll/summary")
def report_payroll_summary():
    """Month-wise payroll summary across all months."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT p.payroll_month,
                       COUNT(*)                AS employees,
                       SUM(p.gross_salary)     AS gross,
                       SUM(p.total_deductions) AS deductions,
                       SUM(p.net_salary)       AS net,
                       SUM(CASE WHEN p.lop_days>0 THEN 1 ELSE 0 END) AS lop_count,
                       MAX(pl.locked_at) IS NOT NULL AND MAX(pl.unlocked_at) IS NULL AS is_locked
                FROM payroll p
                LEFT JOIN payroll_locks pl ON pl.payroll_month=p.payroll_month
                GROUP BY p.payroll_month
                ORDER BY p.payroll_month DESC
            """)
        ).mappings().all()
    return [dict(r) for r in rows]


@app.get("/reports/payroll/department")
def report_department_salary(month: str):
    """Department-wise salary breakdown for a given month."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT department,
                       COUNT(*) AS employees,
                       SUM(gross_salary)     AS gross,
                       SUM(total_deductions) AS deductions,
                       SUM(net_salary)       AS net
                FROM payroll
                WHERE payroll_month=:month
                GROUP BY department
                ORDER BY net DESC
            """),
            {"month": month}
        ).mappings().all()
    return [dict(r) for r in rows]


@app.get("/reports/payroll/employee-history/{emp_id}")
def report_employee_salary_history(emp_id: str):
    """All payslips for a single employee — chronological."""
    with engine.connect() as conn:
        rows = conn.execute(
            text("""
                SELECT payslip_no, payroll_month, gross_salary, total_deductions, net_salary, generated_at
                FROM payslips
                WHERE emp_id=:eid
                ORDER BY payroll_month DESC
            """),
            {"eid": emp_id}
        ).mappings().all()
    return [dict(r) for r in rows]


# ---------- 2. MODELS + ROUTES ----------
 
import json as _json
from datetime import date as _date
 
class ReminderCreate(BaseModel):
    title: str
    description: str = ""
    category: str = "Other"
    priority: str = "Medium"
    assigned_to: int
    created_by: int
    start_date: Optional[str] = None
    due_date: str
    reminder_offsets: List[int] = [7, 3, 1, 0]
    repeat_after_due: bool = False
    amount: Optional[float] = None
    related_module: Optional[str] = None
    related_record_id: Optional[int] = None
    remarks: str = ""
 
 
class ReminderUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None
    assigned_to: Optional[int] = None
    start_date: Optional[str] = None
    due_date: Optional[str] = None
    reminder_offsets: Optional[List[int]] = None
    repeat_after_due: Optional[bool] = None
    amount: Optional[float] = None
    remarks: Optional[str] = None
 
 
class ReminderComplete(BaseModel):
    completed_by: int
    completion_notes: str = ""
 
 
def _days_remaining(due_date):
    if isinstance(due_date, str):
        due_date = _date.fromisoformat(due_date)
    return (due_date - _date.today()).days
 
 
def _reminder_row_to_dict(r):
    d = dict(r)
    try:
        d["reminder_offsets"] = _json.loads(d.get("reminder_offsets_json") or "[]")
    except Exception:
        d["reminder_offsets"] = []
    d["days_remaining"] = _days_remaining(d["due_date"]) if d.get("due_date") else None
    if d.get("amount") is not None:
        d["amount"] = float(d["amount"])
    for k in ("due_date", "start_date", "completed_on", "created_at", "updated_at"):
        if d.get(k) is not None:
            d[k] = str(d[k])
    return d
 
 
# ---------- LIST / SEARCH / FILTER ----------
@app.get("/reminders")
def get_reminders(
    user_id: Optional[int] = None,
    role: Optional[str] = None,
    category: Optional[str] = None,
    priority: Optional[str] = None,
    status: Optional[str] = None,
    assigned_to: Optional[int] = None,
    overdue: Optional[bool] = None,
    search: Optional[str] = None,
):
    filters, params = [], {}
 
    # Role-based visibility: employees only see their own; admins/managers see all
    # Every user only sees reminders they created
    if user_id:
        filters.append("created_by = :uid")
        params["uid"] = user_id
 
    if category:
        filters.append("category = :cat"); params["cat"] = category
    if priority:
        filters.append("priority = :pri"); params["pri"] = priority
    if status:
        filters.append("r.status = :st"); params["st"] = status
    if assigned_to:
        filters.append("r.assigned_to = :ato"); params["ato"] = assigned_to
    if search:
        filters.append("(title ILIKE :s OR description ILIKE :s)")
        params["s"] = f"%{search}%"
    if overdue:
        filters.append("r.due_date < CURRENT_DATE AND r.status = 'Open'")
 
    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"""
                SELECT r.*, u.full_name AS assigned_to_name
                FROM reminders r
                LEFT JOIN users u ON u.id = r.assigned_to
                {where}
                ORDER BY r.due_date ASC
            """),
            params
        ).mappings().all()
 
    return [_reminder_row_to_dict(r) for r in rows]
 
 
# ---------- TODAY'S DUE REMINDERS (for dashboard widget) ----------
@app.get("/reminders/today")
def get_todays_reminders(user_id: int, role: Optional[str] = "employee"):
    """
    Returns reminders for the logged-in user whose days_remaining matches
    one of their configured reminder offsets (or are due/overdue), plus
    any pending travel requests that need this user's attention (all
    pending requests for admins who approve them, or the user's own
    still-pending requests otherwise) so they surface in the same
    dashboard widget / bell / reminders page.
    """

    filters = [
        "r.status = 'Open'",
        "r.created_by = :uid"
    ]

    params = {
        "uid": int(user_id)
    }

    where = "WHERE " + " AND ".join(filters)

    with engine.connect() as conn:
        rows = conn.execute(
            text(f"""
                SELECT r.*
                FROM reminders r
                {where}
                ORDER BY r.due_date ASC
            """),
            params
        ).mappings().all()

        if role == "admin":
            travel_rows = conn.execute(
                text("SELECT * FROM travel_requests WHERE status = 'Pending' ORDER BY created_at ASC")
            ).mappings().all()
        else:
            travel_rows = conn.execute(
                text("SELECT * FROM travel_requests WHERE status = 'Pending' AND user_id = :uid ORDER BY created_at ASC"),
                {"uid": int(user_id)}
            ).mappings().all()

    due_now = []

    for r in rows:
        d = _reminder_row_to_dict(r)

        days = d["days_remaining"]

        offsets = d.get("reminder_offsets") or [7, 3, 1, 0]

        if (days in offsets) or (days <= 0):
            due_now.append(d)

    for tr in travel_rows:
        due_now.append({
            "id": f"travel-{tr['id']}",
            "travel_id": tr["id"],
            "is_travel_request": True,
            "title": (
                f"Travel Request: {tr['employee_name']} → {tr['destination']}"
                if role == "admin" else
                f"Your travel request to {tr['destination']}"
            ),
            "category": "Travel Request",
            "description": tr["purpose"],
            "due_date": str(tr["start_date"]),
            "days_remaining": 0,
            "priority": "High" if role == "admin" else "Medium",
            "amount": float(tr["estimated_cost"]) if tr["estimated_cost"] is not None else None,
            "status": "Open",
            "assigned_to_name": tr["employee_name"],
        })

    due_now.sort(key=lambda x: x["days_remaining"])

    return due_now
 
 
# ---------- NOTIFICATION CENTER FEED ----------
@app.get("/notifications/reminders")
def get_reminder_notifications(user_id: int, role: str = "employee"):
    """Same calculation as /reminders/today, shaped for the notification bell."""
    items = get_todays_reminders(user_id, role)
    feed = []
    for r in items:
        days = r["days_remaining"]
        if days < 0:
            line = f"Overdue by {abs(days)} day(s)"
        elif days == 0:
            line = "Due Today"
        else:
            line = f"Due in {days} day(s)"
        feed.append({
            "id": r["id"],
            "title": r["title"],
            "category": r["category"],
            "priority": r["priority"],
            "line": line,
            "days_remaining": days,
            "due_date": r["due_date"],
        })
    return feed
 
 
# ---------- CREATE ----------
@app.post("/reminders")
def create_reminder(data: ReminderCreate, _staff: dict = Depends(require_roles("admin", "manager"))):
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    INSERT INTO reminders
                    (title, description, category, priority, assigned_to, created_by,
                     start_date, due_date, reminder_offsets_json, repeat_after_due,
                     amount, related_module, related_record_id, remarks, status)
                    VALUES
                    (:title, :description, :category, :priority, :assigned_to, :created_by,
                     :start_date, :due_date, :offsets, :repeat, :amount, :rel_mod, :rel_id,
                     :remarks, 'Open')
                    RETURNING id
                """),
                {
                    "title": data.title,
                    "description": data.description,
                    "category": data.category,
                    "priority": data.priority,
                    "assigned_to": data.assigned_to,
                    "created_by": data.created_by,
                    "start_date": data.start_date,
                    "due_date": data.due_date,
                    "offsets": _json.dumps(data.reminder_offsets),
                    "repeat": data.repeat_after_due,
                    "amount": data.amount,
                    "rel_mod": data.related_module,
                    "rel_id": data.related_record_id,
                    "remarks": data.remarks,
                }
            )
            new_id = result.fetchone()[0]
        return {"success": True, "id": new_id}
    except Exception as e:
        return {"success": False, "message": str(e)}
 
 
# ---------- UPDATE ----------
@app.put("/reminders/{reminder_id}")
def update_reminder(reminder_id: int, data: ReminderUpdate, _staff: dict = Depends(require_roles("admin", "manager"))):
    fields, params = [], {"id": reminder_id}
    payload = data.dict(exclude_unset=True)
 
    if "reminder_offsets" in payload:
        payload["reminder_offsets_json"] = _json.dumps(payload.pop("reminder_offsets"))
 
    for k, v in payload.items():
        fields.append(f"{k} = :{k}")
        params[k] = v
 
    if not fields:
        return {"success": False, "message": "No fields to update."}
 
    fields.append("updated_at = NOW()")
    with engine.begin() as conn:
        result = conn.execute(
            text(f"UPDATE reminders SET {', '.join(fields)} WHERE id = :id RETURNING id"),
            params
        )
        row = result.fetchone()
    if not row:
        return {"success": False, "message": "Reminder not found."}
    return {"success": True}
 
 
# ---------- MARK COMPLETE ----------
@app.post("/reminders/{reminder_id}/complete")
def complete_reminder(reminder_id: int, data: ReminderComplete):
    with engine.begin() as conn:
        result = conn.execute(
            text("""
                UPDATE reminders
                SET status='Completed', completed_on=NOW(),
                    completed_by=:cb, completion_notes=:notes, updated_at=NOW()
                WHERE id=:id
                RETURNING id
            """),
            {"id": reminder_id, "cb": data.completed_by, "notes": data.completion_notes}
        )
        row = result.fetchone()
    if not row:
        return {"success": False, "message": "Reminder not found."}
    return {"success": True}
 
 
# ---------- DELETE ----------
@app.delete("/reminders/{reminder_id}")
def delete_reminder(reminder_id: int, _staff: dict = Depends(require_roles("admin", "manager"))):
    with engine.begin() as conn:
        result = conn.execute(
            text("DELETE FROM reminders WHERE id=:id RETURNING id"),
            {"id": reminder_id}
        )
        row = result.fetchone()
    if not row:
        return {"success": False, "message": "Reminder not found."}
    return {"success": True}


# ── EXCEL EXPORT ──────────────────────────────────────────
import io as _io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook as _Workbook
from openpyxl.styles import Font as _Font, PatternFill as _PatternFill
from fastapi import HTTPException as _HTTPException

EXPORTABLE_TABLES = {
    "users":                     "Users",
    "team_members":              "Team Members",
    "tasks":                     "Tasks",
    "task_history":              "Task History",
    "task_remarks":              "Task Remarks",
    "attendance":                "Attendance",
    "attendance_audit":          "Attendance Audit",
    "attendance_corrections":    "Attendance Corrections",
    "attendance_settings":       "Attendance Settings",
    "correction_timeline":       "Correction Timeline",
    "payroll":                   "Payroll",
    "payroll_locks":             "Payroll Locks",
    "payslips":                  "Payslips",
    "salary_structure":          "Salary Structure",
    "leave_requests":            "Leave Requests",
    "employee_voice":            "Employee Voice",
    "calendar_tasks":            "Calendar Tasks",
    "meetings":                  "Meetings",
    "meeting_attendees":         "Meeting Attendees",
    "reminders":                 "Reminders",
    "holidays":                  "Holidays",
    "holiday_date":              "Holiday Dates",
    "letterheads":               "Letterheads",
    "letterhead_serial_counter": "Letterhead Serial Counter",
    "print_logs":                "Print Logs",
    "insurance_policies":        "Insurance Policies",
    "insurance_claims":          "Insurance Claims",
}

@app.get("/admin/export/{table_name}")
def export_table_excel(table_name: str, _admin: dict = Depends(require_roles("admin"))):
    if table_name not in EXPORTABLE_TABLES:
        raise _HTTPException(status_code=403, detail="Table not allowed for export.")

    try:
        with engine.connect() as conn:
            result = conn.execute(text(f'SELECT * FROM "{table_name}"'))
            rows    = result.fetchall()
            columns = list(result.keys())
    except Exception as e:
        raise _HTTPException(status_code=500, detail=f"Query failed: {str(e)}")

    wb = _Workbook()
    ws = wb.active
    ws.title = EXPORTABLE_TABLES[table_name][:31]

    # Header row with styling
    header_font = _Font(bold=True, color="FFFFFF")
    header_fill = _PatternFill("solid", start_color="1E3A5F")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")

    # Auto-size columns
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0) for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={table_name}_export.xlsx"}
    )


# ════════════════════════════════════════════════════════════════════════
# INSURANCE TRACKING MODULE
# Covers all policy types (Life, Health, Car, Property, Travel, Other).
# Renewal + no-claim-bonus reminders piggyback on the existing `reminders`
# table so they show up in the same bell/widget the rest of the app uses.
# ════════════════════════════════════════════════════════════════════════

import base64 as _b64

# ---------- TABLES (idempotent, safe to run on every startup) ----------
with engine.begin() as conn:
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS insurance_policies (
            id SERIAL PRIMARY KEY,
            policy_type TEXT NOT NULL,                 -- Life, Health, Car, Property, Travel, Other
            policy_number TEXT NOT NULL,
            insurer_name TEXT NOT NULL,
            employee_id INTEGER,
            employee_name TEXT,
            covers_dependents BOOLEAN DEFAULT FALSE,
            dependent_names TEXT,
            vehicle_reg_no TEXT,
            vehicle_details TEXT,
            start_date DATE,
            expiry_date DATE NOT NULL,
            premium_amount NUMERIC,
            premium_frequency TEXT DEFAULT 'Annual',
            sum_insured NUMERIC,
            no_claim_bonus_percent NUMERIC DEFAULT 0,
            cumulative_bonus_percent NUMERIC DEFAULT 0,
            nominee_name TEXT,
            nominee_relation TEXT,
            broker_name TEXT,
            broker_contact TEXT,
            status TEXT DEFAULT 'Active',              -- Active, Expired, Renewed, Lapsed
            file_url TEXT,
            file_name TEXT,
            notes TEXT,
            linked_reminder_id INTEGER,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS insurance_claims (
            id SERIAL PRIMARY KEY,
            policy_id INTEGER REFERENCES insurance_policies(id) ON DELETE CASCADE,
            claim_date DATE NOT NULL,
            claim_amount NUMERIC,
            claim_reason TEXT,
            status TEXT DEFAULT 'Filed',               -- Filed, Approved, Rejected, Settled
            settled_amount NUMERIC,
            notes TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """))


ANTHROPIC_API_KEY = _os.getenv("ANTHROPIC_API_KEY")


# ---------- Pydantic models ----------
class InsurancePolicyIn(BaseModel):
    policy_type: str
    policy_number: str
    insurer_name: str
    employee_id: Optional[int] = None
    employee_name: Optional[str] = None
    covers_dependents: bool = False
    dependent_names: Optional[str] = None
    vehicle_reg_no: Optional[str] = None
    vehicle_details: Optional[str] = None
    start_date: Optional[str] = None
    expiry_date: str
    premium_amount: Optional[float] = None
    premium_frequency: str = "Annual"
    sum_insured: Optional[float] = None
    no_claim_bonus_percent: Optional[float] = 0
    cumulative_bonus_percent: Optional[float] = 0
    nominee_name: Optional[str] = None
    nominee_relation: Optional[str] = None
    broker_name: Optional[str] = None
    broker_contact: Optional[str] = None
    notes: Optional[str] = None
    created_by: int
    role: str  # manual role check -- this app's frontend doesn't send bearer tokens


class InsurancePolicyUpdate(BaseModel):
    policy_type: Optional[str] = None
    policy_number: Optional[str] = None
    insurer_name: Optional[str] = None
    employee_id: Optional[int] = None
    employee_name: Optional[str] = None
    covers_dependents: Optional[bool] = None
    dependent_names: Optional[str] = None
    vehicle_reg_no: Optional[str] = None
    vehicle_details: Optional[str] = None
    start_date: Optional[str] = None
    expiry_date: Optional[str] = None
    premium_amount: Optional[float] = None
    premium_frequency: Optional[str] = None
    sum_insured: Optional[float] = None
    no_claim_bonus_percent: Optional[float] = None
    cumulative_bonus_percent: Optional[float] = None
    nominee_name: Optional[str] = None
    nominee_relation: Optional[str] = None
    broker_name: Optional[str] = None
    broker_contact: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    role: str


class InsuranceRenewIn(BaseModel):
    new_start_date: str
    new_expiry_date: str
    new_premium_amount: Optional[float] = None
    carry_forward_bonus: bool = True
    bonus_increment_percent: Optional[float] = 5
    reset_bonus: bool = False
    created_by: int
    role: str


class InsuranceClaimIn(BaseModel):
    policy_id: int
    claim_date: str
    claim_amount: Optional[float] = None
    claim_reason: Optional[str] = None
    status: str = "Filed"
    settled_amount: Optional[float] = None
    notes: Optional[str] = None
    created_by: int
    role: str


# ---------- Helpers ----------
def _insurance_row_to_dict(r):
    d = dict(r)
    for k in ("start_date", "expiry_date", "created_at", "updated_at"):
        if d.get(k) is not None:
            d[k] = str(d[k])
    if d.get("expiry_date"):
        try:
            d["days_to_expiry"] = (_date.fromisoformat(str(d["expiry_date"])[:10]) - _date.today()).days
        except Exception:
            d["days_to_expiry"] = None
    for k in ("premium_amount", "sum_insured", "no_claim_bonus_percent", "cumulative_bonus_percent"):
        if d.get(k) is not None:
            d[k] = float(d[k])
    return d


def _sync_insurance_reminder(conn, policy_id, expiry_date, policy_number, policy_type,
                              premium_amount, created_by, linked_reminder_id):
    """Create/update the reminders row so renewals surface in the existing bell + widget."""
    title = f"{policy_type} insurance renewal — {policy_number}"
    if linked_reminder_id:
        conn.execute(text("""
            UPDATE reminders SET title=:t, due_date=:d, amount=:a, status='Open', updated_at=NOW()
            WHERE id=:rid
        """), {"t": title, "d": expiry_date, "a": premium_amount, "rid": linked_reminder_id})
        return linked_reminder_id
    result = conn.execute(text("""
        INSERT INTO reminders
            (title, description, category, priority, assigned_to, created_by,
             due_date, reminder_offsets_json, amount, related_module, related_record_id, status)
        VALUES
            (:t, :desc, 'Insurance', 'High', :cb, :cb, :d, :offsets, :a, 'insurance', :pid, 'Open')
        RETURNING id
    """), {
        "t": title, "desc": f"Renewal due for policy {policy_number}.",
        "cb": created_by, "d": expiry_date,
        "offsets": _json.dumps([60, 30, 7, 0]), "a": premium_amount, "pid": policy_id
    })
    return result.fetchone()[0]


# ---------- CREATE ----------
@app.post("/insurance/policies")
def create_insurance_policy(data: InsurancePolicyIn):
    if data.role not in ("admin", "manager"):
        return {"success": False, "message": "You don't have permission to add insurance policies."}
    try:
        payload = data.dict(exclude={"role"})
        with engine.begin() as conn:
            result = conn.execute(text("""
                INSERT INTO insurance_policies
                    (policy_type, policy_number, insurer_name, employee_id, employee_name,
                     covers_dependents, dependent_names, vehicle_reg_no, vehicle_details,
                     start_date, expiry_date, premium_amount, premium_frequency, sum_insured,
                     no_claim_bonus_percent, cumulative_bonus_percent, nominee_name, nominee_relation,
                     broker_name, broker_contact, notes, created_by)
                VALUES
                    (:policy_type,:policy_number,:insurer_name,:employee_id,:employee_name,
                     :covers_dependents,:dependent_names,:vehicle_reg_no,:vehicle_details,
                     :start_date,:expiry_date,:premium_amount,:premium_frequency,:sum_insured,
                     :no_claim_bonus_percent,:cumulative_bonus_percent,:nominee_name,:nominee_relation,
                     :broker_name,:broker_contact,:notes,:created_by)
                RETURNING id
            """), payload)
            new_id = result.fetchone()[0]
            rid = _sync_insurance_reminder(conn, new_id, data.expiry_date, data.policy_number,
                                            data.policy_type, data.premium_amount, data.created_by, None)
            conn.execute(text("UPDATE insurance_policies SET linked_reminder_id=:rid WHERE id=:id"),
                         {"rid": rid, "id": new_id})
        return {"success": True, "id": new_id}
    except Exception as e:
        return {"success": False, "message": str(e)}


# ---------- LIST / SEARCH / FILTER ----------
@app.get("/insurance/policies")
def list_insurance_policies(
    role: str = "employee",
    user_id: Optional[int] = None,
    policy_type: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
):
    filters, params = [], {}
    if role == "employee":
        filters.append("employee_id = :uid")
        params["uid"] = user_id
    if policy_type:
        filters.append("policy_type = :pt"); params["pt"] = policy_type
    if status:
        filters.append("status = :st"); params["st"] = status
    if search:
        filters.append("(policy_number ILIKE :s OR insurer_name ILIKE :s OR employee_name ILIKE :s)")
        params["s"] = f"%{search}%"
    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with engine.connect() as conn:
        rows = conn.execute(
            text(f"SELECT * FROM insurance_policies {where} ORDER BY expiry_date ASC"), params
        ).mappings().all()
    return [_insurance_row_to_dict(r) for r in rows]


# ---------- EXPIRING SOON (for renewal reminder widgets) ----------
@app.get("/insurance/expiring")
def insurance_expiring(days: int = 30):
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT * FROM insurance_policies
            WHERE status = 'Active' AND expiry_date <= CURRENT_DATE + (:days || ' days')::interval
            ORDER BY expiry_date ASC
        """), {"days": days}).mappings().all()
    return [_insurance_row_to_dict(r) for r in rows]


# ---------- DASHBOARD STATS WIDGET ----------
@app.get("/insurance/stats")
def insurance_stats():
    with engine.connect() as conn:
        total = conn.execute(text(
            "SELECT COUNT(*) FROM insurance_policies WHERE status='Active'"
        )).scalar()
        expiring_30 = conn.execute(text("""
            SELECT COUNT(*) FROM insurance_policies
            WHERE status='Active' AND expiry_date <= CURRENT_DATE + INTERVAL '30 days'
        """)).scalar()
        expired = conn.execute(text("""
            SELECT COUNT(*) FROM insurance_policies
            WHERE expiry_date < CURRENT_DATE AND status NOT IN ('Renewed')
        """)).scalar()
        total_premium = conn.execute(text(
            "SELECT COALESCE(SUM(premium_amount),0) FROM insurance_policies WHERE status='Active'"
        )).scalar()
        bonus_at_risk = conn.execute(text("""
            SELECT COUNT(*) FROM insurance_policies
            WHERE status='Active' AND cumulative_bonus_percent > 0
              AND expiry_date <= CURRENT_DATE + INTERVAL '30 days'
        """)).scalar()
        by_type = conn.execute(text("""
            SELECT policy_type, COUNT(*) AS cnt, COALESCE(SUM(premium_amount),0) AS total_premium
            FROM insurance_policies WHERE status='Active' GROUP BY policy_type
        """)).mappings().all()
    return {
        "active_policies": total,
        "expiring_within_30_days": expiring_30,
        "expired": expired,
        "total_annual_premium": float(total_premium),
        "bonus_at_risk_count": bonus_at_risk,
        "by_type": [{"policy_type": r["policy_type"], "count": r["cnt"], "total_premium": float(r["total_premium"])} for r in by_type],
    }


# ---------- GET SINGLE ----------
@app.get("/insurance/policies/{policy_id}")
def get_insurance_policy(policy_id: int):
    with engine.connect() as conn:
        row = conn.execute(text("SELECT * FROM insurance_policies WHERE id=:id"), {"id": policy_id}).mappings().first()
    if not row:
        return {"success": False, "message": "Policy not found."}
    return _insurance_row_to_dict(row)


# ---------- UPDATE ----------
@app.put("/insurance/policies/{policy_id}")
def update_insurance_policy(policy_id: int, data: InsurancePolicyUpdate):
    if data.role not in ("admin", "manager"):
        return {"success": False, "message": "You don't have permission to edit insurance policies."}
    payload = data.dict(exclude={"role"}, exclude_unset=True)
    if not payload:
        return {"success": False, "message": "No fields to update."}
    fields, params = [], {"id": policy_id}
    for k, v in payload.items():
        fields.append(f"{k} = :{k}")
        params[k] = v
    fields.append("updated_at = NOW()")
    with engine.begin() as conn:
        row = conn.execute(
            text(f"UPDATE insurance_policies SET {', '.join(fields)} WHERE id=:id RETURNING *"), params
        ).mappings().first()
        if row and "expiry_date" in payload:
            _sync_insurance_reminder(conn, policy_id, row["expiry_date"], row["policy_number"],
                                      row["policy_type"], row["premium_amount"],
                                      row["created_by"] or 0, row["linked_reminder_id"])
    if not row:
        return {"success": False, "message": "Policy not found."}
    return {"success": True}


# ---------- RENEW (carries forward / resets no-claim bonus) ----------
@app.post("/insurance/policies/{policy_id}/renew")
def renew_insurance_policy(policy_id: int, data: InsuranceRenewIn):
    if data.role not in ("admin", "manager"):
        return {"success": False, "message": "You don't have permission to renew insurance policies."}
    with engine.begin() as conn:
        row = conn.execute(text("SELECT * FROM insurance_policies WHERE id=:id"), {"id": policy_id}).mappings().first()
        if not row:
            return {"success": False, "message": "Policy not found."}

        claims_count = conn.execute(text("""
            SELECT COUNT(*) FROM insurance_claims WHERE policy_id=:id AND claim_date >= :sd
        """), {"id": policy_id, "sd": row["start_date"]}).scalar()

        if data.reset_bonus or claims_count > 0:
            new_bonus = 0
        elif data.carry_forward_bonus:
            new_bonus = float(row["cumulative_bonus_percent"] or 0) + float(data.bonus_increment_percent or 0)
        else:
            new_bonus = float(row["cumulative_bonus_percent"] or 0)

        new_premium = data.new_premium_amount if data.new_premium_amount is not None else row["premium_amount"]

        conn.execute(text("""
            UPDATE insurance_policies
            SET start_date=:sd, expiry_date=:ed, premium_amount=:pa,
                cumulative_bonus_percent=:bonus, status='Active', updated_at=NOW()
            WHERE id=:id
        """), {"sd": data.new_start_date, "ed": data.new_expiry_date, "pa": new_premium,
               "bonus": new_bonus, "id": policy_id})

        _sync_insurance_reminder(conn, policy_id, data.new_expiry_date, row["policy_number"],
                                  row["policy_type"], new_premium, data.created_by, row["linked_reminder_id"])

    return {"success": True, "claims_since_last_cycle": claims_count, "new_bonus_percent": new_bonus}


# ---------- DELETE ----------
@app.delete("/insurance/policies/{policy_id}")
def delete_insurance_policy(policy_id: int, role: str = "employee"):
    if role not in ("admin", "manager"):
        return {"success": False, "message": "You don't have permission to delete insurance policies."}
    with engine.begin() as conn:
        row = conn.execute(text("SELECT linked_reminder_id FROM insurance_policies WHERE id=:id"),
                            {"id": policy_id}).mappings().first()
        if not row:
            return {"success": False, "message": "Policy not found."}
        conn.execute(text("DELETE FROM insurance_policies WHERE id=:id"), {"id": policy_id})
        if row["linked_reminder_id"]:
            conn.execute(text("DELETE FROM reminders WHERE id=:rid"), {"rid": row["linked_reminder_id"]})
    return {"success": True}


# ---------- CLAIMS ----------
@app.post("/insurance/claims")
def add_insurance_claim(data: InsuranceClaimIn):
    if data.role not in ("admin", "manager"):
        return {"success": False, "message": "You don't have permission to log claims."}
    try:
        payload = data.dict(exclude={"role"})
        with engine.begin() as conn:
            result = conn.execute(text("""
                INSERT INTO insurance_claims
                    (policy_id, claim_date, claim_amount, claim_reason, status, settled_amount, notes, created_by)
                VALUES
                    (:policy_id,:claim_date,:claim_amount,:claim_reason,:status,:settled_amount,:notes,:created_by)
                RETURNING id
            """), payload)
            new_id = result.fetchone()[0]
        return {"success": True, "id": new_id}
    except Exception as e:
        return {"success": False, "message": str(e)}


@app.get("/insurance/policies/{policy_id}/claims")
def get_policy_claims(policy_id: int):
    with engine.connect() as conn:
        rows = conn.execute(
            text("SELECT * FROM insurance_claims WHERE policy_id=:id ORDER BY claim_date DESC"),
            {"id": policy_id}
        ).mappings().all()
    out = []
    for r in rows:
        d = dict(r)
        for k in ("claim_date", "created_at"):
            if d.get(k) is not None:
                d[k] = str(d[k])
        for k in ("claim_amount", "settled_amount"):
            if d.get(k) is not None:
                d[k] = float(d[k])
        out.append(d)
    return out


# ---------- POLICY DOCUMENT UPLOAD / SERVE ----------
@app.post("/insurance/policies/{policy_id}/upload-document")
async def upload_insurance_document(policy_id: int, file: UploadFile = FastAPIFile(...)):
    safe_name = f"{policy_id}_{int(_auth_time.time())}_{file.filename.replace(' ', '_')}"
    contents  = await file.read()
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO stored_files (filename, content_type, data, original_name)
                VALUES (:fn, :ct, :data, :orig)
                ON CONFLICT (filename) DO UPDATE
                SET content_type = EXCLUDED.content_type,
                    data = EXCLUDED.data,
                    original_name = EXCLUDED.original_name,
                    created_at = now()
            """),
            {"fn": safe_name, "ct": file.content_type, "data": contents, "orig": file.filename}
        )
    file_url = f"/insurance/document/{safe_name}"
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE insurance_policies SET file_url=:u, file_name=:n WHERE id=:id"),
            {"u": file_url, "n": file.filename, "id": policy_id}
        )
    return {"success": True, "file_url": file_url, "file_name": file.filename}


@app.get("/insurance/document/{filename}")
def serve_insurance_document(filename: str):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT content_type, data, original_name FROM stored_files WHERE filename=:fn"),
            {"fn": filename}
        ).mappings().first()
    if not row:
        return {"error": "File not found"}
    return _RawResponse(
        content=bytes(row["data"]),
        media_type=row["content_type"] or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{row["original_name"] or filename}"'}
    )


# ---------- AI DOCUMENT EXTRACTION (pre-fill the Add Policy form) ----------
@app.post("/insurance/extract-document")
async def extract_insurance_document(file: UploadFile = FastAPIFile(...)):
    """
    Reads an uploaded policy document (PDF or image) and asks Claude to pull out
    structured fields. Returns a best-effort JSON the frontend uses to pre-fill
    the Add Policy form -- the user should still review before saving.
    """
    if not ANTHROPIC_API_KEY:
        return {"success": False, "message": "Document extraction isn't configured yet. Set ANTHROPIC_API_KEY on the server."}

    raw = await file.read()
    b64data = _b64.b64encode(raw).decode()
    ext = (file.filename or "").lower().split(".")[-1]
    if ext == "pdf":
        block_type, media_type = "document", "application/pdf"
    else:
        block_type = "image"
        media_type = f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"

    prompt = (
        "You are reading an insurance policy document. Extract the following fields and reply with "
        "ONLY a strict JSON object -- no markdown fences, no commentary. Use null for anything not present.\n"
        "{\n"
        '  "policy_type": one of "Life", "Health", "Car", "Property", "Travel", "Other",\n'
        '  "policy_number": string,\n'
        '  "insurer_name": string,\n'
        '  "start_date": "YYYY-MM-DD",\n'
        '  "expiry_date": "YYYY-MM-DD",\n'
        '  "premium_amount": number,\n'
        '  "sum_insured": number,\n'
        '  "no_claim_bonus_percent": number,\n'
        '  "nominee_name": string,\n'
        '  "nominee_relation": string,\n'
        '  "vehicle_reg_no": string\n'
        "}"
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-5",
                "max_tokens": 1000,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": block_type, "source": {"type": "base64", "media_type": media_type, "data": b64data}},
                        {"type": "text", "text": prompt},
                    ],
                }],
            },
            timeout=60,
        )
        result = resp.json()
        if "content" not in result:
            return {"success": False, "message": result.get("error", {}).get("message", "Extraction failed.")}
        text_out = "".join(b.get("text", "") for b in result.get("content", []) if b.get("type") == "text")
        cleaned = text_out.replace("```json", "").replace("```", "").strip()
        extracted = _json.loads(cleaned)
        return {"success": True, "data": extracted}
    except Exception as e:
        return {"success": False, "message": f"Couldn't read that document automatically: {e}"}
