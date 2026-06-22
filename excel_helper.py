# excel_helper.py

```python
"""
excel_helper.py
────────────────────────────
Excel backend for Panache Sales Tracker

Workbook:
sales_tracker.xlsx

Sheets:
Employee_{user_id}
ChangeRequests
"""

import os
import uuid
import datetime
from openpyxl import Workbook, load_workbook

FILE_NAME = "sales_tracker.xlsx"

SALES_HEADERS = [
    "week",
    "customer",
    "product",
    "projected",
    "price",
    "achieved"
]

CR_HEADERS = [
    "request_id",
    "employee_id",
    "week",
    "customer",
    "product",
    "old_qty",
    "new_qty",
    "reason",
    "status",
    "manager_note",
    "created_at"
]


# ─────────────────────────────────────────────
# WORKBOOK
# ─────────────────────────────────────────────

def _get_workbook():

    if not os.path.exists(FILE_NAME):

        wb = Workbook()

        ws = wb.active
        ws.title = "ChangeRequests"
        ws.append(CR_HEADERS)

        wb.save(FILE_NAME)

    return load_workbook(FILE_NAME)


# ─────────────────────────────────────────────
# EMPLOYEE SHEETS
# ─────────────────────────────────────────────

def _employee_sheet(user_id):

    wb = _get_workbook()

    sheet_name = f"Employee_{user_id}"

    if sheet_name not in wb.sheetnames:

        ws = wb.create_sheet(sheet_name)
        ws.append(SALES_HEADERS)

        wb.save(FILE_NAME)

    return wb, wb[sheet_name]


# ─────────────────────────────────────────────
# CHANGE REQUEST SHEET
# ─────────────────────────────────────────────

def _cr_sheet():

    wb = _get_workbook()

    ws = wb["ChangeRequests"]

    return wb, ws


# ─────────────────────────────────────────────
# WEEK
# ─────────────────────────────────────────────

def current_week():

    return datetime.date.today().isocalendar()[1]


# ─────────────────────────────────────────────
# INTERNAL HELPERS
# ─────────────────────────────────────────────

def _find_row(ws, week, customer, product):

    for row in ws.iter_rows(min_row=2):

        if (
            str(row[0].value) == str(week)
            and str(row[1].value).strip().lower() == customer.strip().lower()
            and str(row[2].value).strip().lower() == product.strip().lower()
        ):
            return row

    return None


# ─────────────────────────────────────────────
# SALES
# ─────────────────────────────────────────────

def add_projection(
    user_id,
    week,
    customer,
    product,
    projected,
    price
):

    wb, ws = _employee_sheet(user_id)

    if _find_row(ws, week, customer, product):

        raise ValueError(
            "Projection already exists"
        )

    ws.append([
        week,
        customer,
        product,
        projected,
        price,
        0
    ])

    wb.save(FILE_NAME)

    return {"success": True}


def get_sales(user_id, week):

    wb, ws = _employee_sheet(user_id)

    data = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if str(row[0]) == str(week):

            data.append({

                "week": row[0],
                "customer": row[1],
                "product": row[2],
                "projected": row[3],
                "price": row[4],
                "achieved": row[5]

            })

    return data


def update_achieved(
    user_id,
    week,
    customer,
    product,
    achieved
):

    wb, ws = _employee_sheet(user_id)

    row = _find_row(
        ws,
        week,
        customer,
        product
    )

    if not row:

        raise ValueError("Row not found")

    row[5].value = achieved

    wb.save(FILE_NAME)

    return {"success": True}


def get_all_weeks(user_id):

    wb, ws = _employee_sheet(user_id)

    weeks = set()

    for row in ws.iter_rows(min_row=2):

        if row[0].value:
            weeks.add(int(row[0].value))

    return sorted(list(weeks))


# ─────────────────────────────────────────────
# MANAGER / ADMIN VIEW
# ─────────────────────────────────────────────

def get_sales_for_employees(
    employee_ids,
    week
):

    result = {}

    for emp in employee_ids:

        try:

            result[emp] = get_sales(
                emp,
                week
            )

        except Exception:

            result[emp] = []

    return result


# ─────────────────────────────────────────────
# MANAGER / ADMIN UPDATE PROJECTION
# ─────────────────────────────────────────────

def admin_update_projection(
    user_id,
    week,
    customer,
    product,
    new_qty
):

    wb, ws = _employee_sheet(user_id)

    row = _find_row(
        ws,
        week,
        customer,
        product
    )

    if not row:

        raise ValueError(
            "Row not found"
        )

    row[3].value = new_qty

    wb.save(FILE_NAME)

    return {"success": True}


# ─────────────────────────────────────────────
# CHANGE REQUESTS
# ─────────────────────────────────────────────

def raise_change_request(
    employee_id,
    week,
    customer,
    product,
    old_qty,
    new_qty,
    reason
):

    wb, ws = _cr_sheet()

    rid = str(uuid.uuid4())

    ws.append([

        rid,
        employee_id,
        week,
        customer,
        product,
        old_qty,
        new_qty,
        reason,
        "pending",
        "",
        datetime.datetime.utcnow().isoformat()

    ])

    wb.save(FILE_NAME)

    return {

        "success": True,
        "request_id": rid

    }


def get_change_requests(
    employee_ids=None,
    status=None
):

    wb, ws = _cr_sheet()

    data = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        record = {

            "request_id": row[0],
            "employee_id": row[1],
            "week": row[2],
            "customer": row[3],
            "product": row[4],
            "old_qty": row[5],
            "new_qty": row[6],
            "reason": row[7],
            "status": row[8],
            "manager_note": row[9],
            "created_at": row[10]

        }

        if employee_ids and record["employee_id"] not in employee_ids:
            continue

        if status and record["status"] != status:
            continue

        data.append(record)

    return data


def resolve_change_request(
    request_id,
    action,
    manager_note=""
):

    wb, ws = _cr_sheet()

    for row in ws.iter_rows(min_row=2):

        if row[0].value == request_id:

            row[8].value = action
            row[9].value = manager_note

            if action == "approved":

                admin_update_projection(

                    row[1].value,
                    row[2].value,
                    row[3].value,
                    row[4].value,
                    row[6].value

                )

            wb.save(FILE_NAME)

            return {"success": True}

    raise ValueError(
        "Request not found"
    )
```
